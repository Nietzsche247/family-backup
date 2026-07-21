import sys, os, shutil
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map, set_literal
from openpyxl import load_workbook
M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'; WB=M+r'\Wilson.xlsx'; SC=M+r'\Wilson.schema.json'
assert not os.path.exists(M+r'\~$Wilson.xlsx'), "OPEN in Excel"
bk=M+r'\Wilson_prewrite_backup_2026-07-20_CD19_MTGSRC.xlsx'; shutil.copy2(WB,bk); print('BACKUP',bk)
s=load_json_strict(SC); wb=load_workbook(WB); cl=wb['CLAIMS & DEFENSES']; cc=col_map(s,'CLAIMS & DEFENSES'); nc=cc['Notes']
assert cl.cell(row=24,column=cc['ID']).value=='CD-0019'
add=(' 5/19 MEETING PRIMARY SOURCE (Gemini 2026-07-20): NO formal transcript exists; documented in the "Christine & Ownership" Google Chat room (source 47) where '
 'Michael Baker noted verbatim: "Meeting with Dave Wilson (At Client\'s house) Scott & Michael, Monday, May 19, 2025, 9:30 - 10:00am." '
 'CONTEMPORANEOUS ATTENDEES = Michael + Scott. This resolves the attendee variance: Wilson\'s 2026-07-19 recollection of Michael+Scott+Martin is not matched by the '
 'contemporaneous note; Wilson separately "approached Martin one last time" per his own 2026-07-19 email, which likely bleeds into his memory. Scheduling + the 5/16 push '
 '("in bed sick") also in Christine & Ownership; broader ops in the "Managers" room. TO PRESERVE via Google Vault or Takeout export of both rooms: '
 'Christine & Ownership = room AAAALofXX3I; Managers = room AAAA5WGJYgc. Once exported + hashed, build the formal MASTER CHRONOLOGY event for the 2025-05-19 meeting.')
set_literal(cl.cell(row=24,column=nc),(str(cl.cell(row=24,column=nc).value or '').rstrip()+add).strip())
wb.save(WB); print('CD-0019 updated + SAVED')
