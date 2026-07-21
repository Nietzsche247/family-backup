import sys, os, shutil
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map, set_literal
from openpyxl import load_workbook
M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'; WB=M+r'\Wilson.xlsx'; SC=M+r'\Wilson.schema.json'
assert not os.path.exists(M+r'\~$Wilson.xlsx'), "OPEN in Excel"
bk=M+r'\Wilson_prewrite_backup_2026-07-21_CONSTREVT.xlsx'; shutil.copy2(WB,bk); print('BACKUP',bk)
s=load_json_strict(SC); wb=load_workbook(WB); ws=wb['MASTER CHRONOLOGY']; c=col_map(s,'MASTER CHRONOLOGY')
CONST={'Transcription State':'NOT CHECKED','Source Location State':'PRESERVED','Authenticity / Foundation':'NOT ASSESSED',
 'Admissibility':'NOT ASSESSED','Counsel Status':'UNREVIEWED','Entered By':'aaron baker','Entered On':'2026-07-21','Record State':'DRAFT'}
EV=[
 dict(id='EVT-0031',date='2025-04-08',dt='EVENT',sc='SYSTEM RECORD',attr='Buildertrend Daily Logs',act='ACT-0002; ACT-0007',src='SRC-0120',tag='TERM',hot='Y',corr='SINGLE SOURCE',
   text="Tom Delaney's last daily log (tree/palm selections). Tom authored only early logs (2025-02-26 sewer relocation, 2025-04-03 site prep, 2025-04-08) and drops out of construction about a month before his 2025-05-09 termination. Scott Culver is the dominant PM author for the actual build.",
   loc="Daily Logs; Tom Delaney last authored 2025-04-08. Site prep complete 2025-04-03 per Tom's log."),
 dict(id='EVT-0032',date='2025-05-19',dt='EVENT',sc='SYSTEM RECORD',attr='Buildertrend Schedule + Daily Logs',act='ACT-0007',src='SRC-0119; SRC-0120',tag='DOC',hot='Y',corr='MULTIPLE INDEPENDENT',
   text="Daily logs run dense and continuous through Tom's 2025-05-09 termination and the 2025-05-19 warning meeting, with no work stoppage. Milestones in the window: pool/pond excavation done ~May 3-5, plumbing complete May 8, steel/rebar May 12-16, ramada gas/electric lines by Omni crew May 14-16, mister trenching May 19, gunite/shotcrete May 28. Scott Culver PM.",
   loc='Schedule tasks + Daily Logs, Apr 29 - May 28, 2025.'),
 dict(id='EVT-0033',date='2025-08-20',dt='EVENT',sc='SYSTEM RECORD',attr='Buildertrend Schedule',act='ACT-0001',src='SRC-0119',tag='CHG',hot='',corr='SINGLE SOURCE',
   text="Ramada footers, inspection, and posts built by Schneider (Jun 6 - Aug 20, 2025), the schedule's only ramada task and explicitly assigned to Schneider - matching CO-0036 (approved 2025-05-24). Omni's crew had run the ramada gas/electric lines May 14-16 before the handoff.",
   loc="Schedule task 'Schneider: Ramada Footers, Inspect, Posts', Jun 6 - Aug 20 2025. Ties to CO-0036 / TXN-0013."),
 dict(id='EVT-0034',date='2025-11-03',dt='EVENT',sc='SYSTEM RECORD',attr='Buildertrend Schedule',act='ACT-0007',src='SRC-0119',tag='DOC',hot='Y',corr='SINGLE SOURCE',
   text='Final Inspection PASSED 2025-11-03.',loc="Schedule task 'Final Inspection' 2025-11-03, marked PASSED."),
 dict(id='EVT-0035',date='2025-11-05',dt='EVENT',sc='SYSTEM RECORD',attr='Buildertrend Schedule',act='ACT-0001',src='SRC-0119',tag='DEF',hot='Y',corr='SINGLE SOURCE',
   text="Pool interior finish (pebble) installed by MMG-Interior Install 2025-11-05/06, polish/detail Nov 7. This interior finish is the subject of Wilson's later ROC complaint #2026-00144 (EVT-0027) and his 2026 'pebble-shine finish' complaint (EVT-0030).",
   loc="Schedule task 'MMG-Interior Install' Nov 5-6 2025. Subject of ROC 2026-00144."),
 dict(id='EVT-0036',date='2025-11-10',dt='EVENT',sc='SYSTEM RECORD',attr='Buildertrend Schedule + Daily Logs',act='ACT-0001',src='SRC-0119; SRC-0120',tag='DOC',hot='Y',corr='MULTIPLE INDEPENDENT',
   text="Pool Start-Up 2025-11-10; water in Nov 8. Job substantially complete. Only the punch list / 5% retention (invoice 0007) remains. 2026 daily logs are warranty/service tickets that repeat 'Start up date: 11/10/25'.",
   loc="Schedule 'Pool Start-Up' 2025-11-10; Daily Logs confirm startup date 11/10/25."),
 dict(id='EVT-0037',date='2026-07-22',dt='EVENT',sc='SYSTEM RECORD',attr='Buildertrend Schedule',act='ACT-0001',src='SRC-0119',tag='WARR',hot='Y',corr='SINGLE SOURCE',
   text="Final Walkthrough scheduled 2026-07-22 and Punch List 2026-07-23 to 08-05 (Phase 3 Interior to Punch-List runs to 08-13), not yet performed (0%). This outstanding punch-list work is tied to invoice 0007, the 5% retention ($51,752.62), and coincides with the current 2026 warranty dispute.",
   loc='Schedule tasks: Final Walkthrough 2026-07-22 (0%), Punch List 2026-07-23 to 08-05. Ties to invoice 0007.'),
]
assert ws.cell(row=35,column=1).value=='EVT-0030'
def P(r,k,v): set_literal(ws.cell(row=r,column=c[k]),v)
for i,e in enumerate(EV):
    r=36+i
    assert not ws.cell(row=r,column=1).value, 'row %d not empty'%r
    P(r,'Event ID',e['id']); P(r,'Event Date',e['date']); P(r,'Date Type',e['dt'])
    P(r,'Event Text (released core)',e['text']); P(r,'Statement Class',e['sc']); P(r,'Attribution / Speaker',e['attr'])
    P(r,'Actors',e['act']); P(r,'Source ID(s)',e['src']); P(r,'Locator',e['loc']); P(r,'Issue Tags',e['tag'])
    if e['hot']: P(r,'Hot','Y')
    P(r,'Dispute Status','UNDISPUTED FOR NOW'); P(r,'Corroboration',e['corr'])
    for k,v in CONST.items(): P(r,k,v)
    P(r,'Batch / Ordinal','HAND-ENTRY-2026-07-21/%d'%(i+1))
    ws.cell(row=r,column=c['Check (auto)']).value='=IF($A%d="","",TRIM(IF($B%d="","NO DATE ","")&IF($J%d="","NO SOURCE ","")&IF($L%d="","NO TAGS ","")))'%(r,r,r,r)
    ws.cell(row=r,column=c['HotRank (auto)']).value='=IF($M%d="Y",COUNTIF($M$6:$M%d,"Y"),"")'%(r,r)
    print('added',e['id'],'row',r)
wb.save(WB); print('SAVED')
