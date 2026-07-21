import sys, os, shutil, hashlib
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map, set_literal
from openpyxl import load_workbook
M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'; WB=M+r'\Wilson.xlsx'; SC=M+r'\Wilson.schema.json'
KEY=M+r'\preserved\keydocs'; VE=M+r'\vault_export'
assert not os.path.exists(M+r'\~$Wilson.xlsx'), "OPEN in Excel"
h130=hashlib.sha256(open(os.path.join(VE,'WILSON_chat_2026-07-21-1.zip'),'rb').read()).hexdigest()
h131=hashlib.sha256(open(os.path.join(KEY,'SRC-0131_2025-05_Tom-termination-Wilson-warning-chat.txt'),'rb').read()).hexdigest()
bk=M+r'\Wilson_prewrite_backup_2026-07-21_VAULT.xlsx'; shutil.copy2(WB,bk); print('BACKUP',bk)
s=load_json_strict(SC); wb=load_workbook(WB)
si=wb['SOURCE INDEX']; c=col_map(s,'SOURCE INDEX')
r0=None
for rr in range(6,si.max_row+2):
    if not si.cell(row=rr,column=c['Source ID']).value: r0=rr; break
assert r0==33,'expected 33 got %s'%r0
SRCS={33:{'Source ID':'SRC-0130','SHA-256':h130,'Doc Date':'2026-07-21','Date Type':'MINED',
  'Date Basis':'Google Vault export generated 2026-07-21; data spans both rooms full history (Dec 2023 - Jul 2026).','Type':'Chat',
  'Description':'Google Vault Chat export of "Christine & Ownership" (AAAALofXX3I) and "Managers" (AAAA5WGJYgc) spaces, full history. 510 messages, MBOX, 1.3 GB (attachments included).',
  'Custodian':'Omni (Google Workspace) via Google Vault','From / Author':'Google Vault export (admin)','To / Recipients':'',
  'Native (relative URI)':'Google Vault matter OMNI-LIT-2026-001, export WILSON_chat_2026-07-21 (MBOX)',
  'Working Copy (filename)':'vault_export/WILSON_chat_2026-07-21-1.zip (+ metadata.xml, errors.csv, .md5)',
  'Acquisition Method':'Exported from Google Vault (matter OMNI-LIT-2026-001, counsel-directed) 2026-07-21; all 4 export files downloaded. PRESERVED SEPARATELY (1.3 GB) - NOT in the emailable bundle due to size; available on request.',
  'Received Date':'2026-07-21','Disposition':'PRESERVED',
  'Notes':'Master Chat export. Google MD5 (zip): 13aa95fcbd9a70c4b9003d4cfc31fe27. Full set in vault_export/ (outside bundle). SRC-0131 = extracted Tom-termination / Wilson-warning-meeting messages (in bundle).'},
 34:{'Source ID':'SRC-0131','SHA-256':h131,'Doc Date':'2025-05-16','Date Type':'DOCUMENT',
  'Date Basis':'Extracted chat from SRC-0130; messages span 2025-05-08 to 2025-05-20 (Tom-termination + Wilson-warning-meeting period).','Type':'Chat',
  'Description':'Extract from SRC-0130: the Tom-termination and Wilson-warning-meeting messages, 2025-05-08 to 2025-05-20, both rooms. Rule-based filter (date window + Tom/Wilson/meeting/termination terms), text.',
  'Custodian':'Omni (Google Workspace) via Google Vault','From / Author':'Omni team (Michael, Scott, Christine, Martin, Aaron, Veronica)','To / Recipients':'internal (Christine & Ownership / Managers rooms)',
  'Native (relative URI)':'derived from SRC-0130 (WILSON_chat_2026-07-21 MBOX)','Working Copy (filename)':'SRC-0131_2025-05_Tom-termination-Wilson-warning-chat.txt',
  'Acquisition Method':'Text extracted + base64 attachments stripped from the Vault MBOX (SRC-0130) 2026-07-21; date+keyword filtered to the warning-meeting period.',
  'Received Date':'2026-07-21','Disposition':'PRESERVED',
  'Notes':'Convenience extract for EVT-0041 (2025-05-19 warning meeting). Authoritative source = SRC-0130. Key lines: Michael "michael is meeting him within the next hour"; "The dave meeting was rescheduled to Monday. He is in bed sick" (Wilson meeting Fri 5/16 -> Mon 5/19).'}}
for rr,d in SRCS.items():
    for k,v in d.items(): set_literal(si.cell(row=rr,column=c[k]),v)
    print('src',d['Source ID'],'row',rr)
# EVT-0041
ch=wb['MASTER CHRONOLOGY']; cc=col_map(s,'MASTER CHRONOLOGY')
er=None
for rr in range(6,ch.max_row+2):
    if not ch.cell(row=rr,column=1).value: er=rr; break
assert er==46,'expected 46 got %s'%er
COMMON={'Witness (disclosure)':None,'Transcription State':'NOT CHECKED','Source Location State':'PRESERVED',
 'Authenticity / Foundation':'NOT ASSESSED','Admissibility':'NOT ASSESSED','Counsel Status':'UNREVIEWED',
 'Entered By':'aaron baker','Entered On':'2026-07-21','Record State':'DRAFT','Dispute Status':'UNDISPUTED FOR NOW'}
d={'Event ID':'EVT-0041','Event Date':'2025-05-19','Date Type':'EVENT','Qualifier':'on or about',
 'Event Text (released core)':"Omni met with Wilson at his home to warn him Tom Delaney was terminated (effective 2025-05-09) and no longer authorized to represent Omni. Prompted by Tom appearing at the Wilson job in Omni apparel; the meeting was set for Fri 2025-05-16 but rescheduled to Monday 2025-05-19 because Wilson was ill. Wilson later confirmed it (SRC-0118): 'you, Scott and Martin came to my house and cornered me regarding Tom's termination and criminal behavior.'",
 'Statement Class':'OBSERVED EVENT','Attribution / Speaker':'Omni (Michael Baker et al.); corroborated by Wilson','Actors':'ACT-0005; ACT-0007; ACT-0001; ACT-0006',
 'Source ID(s)':'SRC-0131; SRC-0118','Locator':'SRC-0131 (Managers 2025-05-16: "michael is meeting him within the next hour"; "The dave meeting was rescheduled to Monday. He is in bed sick"); SRC-0118 (Wilson 2026-07-19 admission). Full export = SRC-0130.',
 'Issue Tags':'TERM','Hot':'Y','Corroboration':'MULTIPLE INDEPENDENT','Batch / Ordinal':'VAULT-CHAT-2026-07-21/1'}
full=dict(COMMON); full.update(d)
print('E46 len =',len(d['Event Text (released core)']))
for k,v in full.items():
    if v is not None: set_literal(ch.cell(row=er,column=cc[k]),v)
ch.cell(row=er,column=cc['Check (auto)']).value='=IF($A%d="","",TRIM(IF($B%d="","NO DATE ","")&IF($J%d="","NO SOURCE ","")&IF($L%d="","NO TAGS ","")))'%(er,er,er,er)
ch.cell(row=er,column=cc['HotRank (auto)']).value='=IF($M%d="Y",COUNTIF($M$6:$M%d,"Y"),"")'%(er,er)
print('EVT-0041 row',er)
wb.save(WB); print('SAVED')
