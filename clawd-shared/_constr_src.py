import sys, os, shutil, hashlib
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map, set_literal
from openpyxl import load_workbook
M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'; WB=M+r'\Wilson.xlsx'; SC=M+r'\Wilson.schema.json'
KEY=M+r'\preserved\keydocs'; DL=r'C:\Users\aaron\Downloads'
assert not os.path.exists(M+r'\~$Wilson.xlsx'), "OPEN in Excel"
def cp(src,dn):
    d=os.path.join(KEY,dn); shutil.copy2(os.path.join(DL,src),d)
    return hashlib.sha256(open(d,'rb').read()).hexdigest()
h119=cp('Schedule_List_Wilson, David.xlsx','SRC-0119_schedule_WILSON.xlsx')
h120=cp('Daily Log Print.pdf','SRC-0120_daily-logs_WILSON.pdf')
print('SRC-0119',h119); print('SRC-0120',h120)
bk=M+r'\Wilson_prewrite_backup_2026-07-21_CONSTRSRC.xlsx'; shutil.copy2(WB,bk); print('BACKUP',bk)
s=load_json_strict(SC); wb=load_workbook(WB); si=wb['SOURCE INDEX']; c=col_map(s,'SOURCE INDEX')
r=None
for rr in range(6,si.max_row+2):
    if not si.cell(row=rr,column=c['Source ID']).value: r=rr; break
assert r==22,'expected row 22 got %s'%r
def P(row,k,v): set_literal(si.cell(row=row,column=c[k]),v)
d119={'Source ID':'SRC-0119','SHA-256':h119,'Doc Date':'2026-07-21','Date Type':'CAPTURE',
 'Date Basis':'Export generated 2026-07-21 (Buildertrend Schedule export).','Type':'Buildertrend',
 'Description':'Buildertrend Schedule export - Wilson job 38668322 (150 tasks with start/end dates; construction timeline)',
 'Custodian':'Omni Pool Builders (Buildertrend job 38668322)','Native (relative URI)':'dir:staging_construction/Schedule_WILSON_2026-07-21.xlsx',
 'Working Copy (filename)':'SRC-0119_schedule_WILSON.xlsx',
 'Acquisition Method':'A1 administrative export; Buildertrend Schedule Export to Excel (POST /apix/v2/Schedules/export 200); Wilson job 38668322; exported 2026-07-21. Downloaded as "Schedule_List_Wilson, David.xlsx".',
 'Received Date':'2026-07-21','Disposition':'PRESERVED',
 'Notes':'Milestones: excavation Apr 29-May 5 2025; plumbing May 6-8; steel May 12-16; gunite/shotcrete May 28; pool tile May 30; decking Jun 4-Oct 30; equipment/electrical final Oct 31; Final Inspection PASSED Nov 3; interior/plaster (MMG-Interior Install) Nov 5-6; fill Nov 8; startup Nov 10 2025. Ramada (Schneider) Jun 6-Aug 20 (matches CO-0036). Landscaping task absent (matches CO-0048 removal). Final Walkthrough planned Jul 22 2026; Punch List to Aug 2026 = invoice 0007 5% retention. Feeds EVT-0031..0037.'}
d120={'Source ID':'SRC-0120','SHA-256':h120,'Doc Date':'2026-07-21','Date Type':'CAPTURE',
 'Date Basis':'Print-to-PDF captured 2026-07-21 (Buildertrend Daily Logs list; log content spans 2025-02-11 to 2026-07-20).','Type':'Buildertrend',
 'Description':'Buildertrend Daily Logs list (401 logs, 2025-02-11 to 2026-07-20) - Wilson job',
 'Custodian':'Omni Pool Builders (Buildertrend job 38668322)','Native (relative URI)':'dir:staging_construction/DailyLogs_WILSON_2026-07-21.pdf',
 'Working Copy (filename)':'SRC-0120_daily-logs_WILSON.pdf',
 'Acquisition Method':'Buildertrend Daily Logs List print view saved to PDF (native print dialog), 2026-07-21; Wilson job 38668322. Downloaded as "Daily Log Print.pdf".',
 'Received Date':'2026-07-21','Disposition':'PRESERVED',
 'Notes':'First log 2025-02-11 (Allysa Redmon, setup). Last log 2026-07-20 (Michael Baker, warranties/final payment). Site prep complete 2025-04-03 (Tom Delaney: "All site prep is complete and ready for construction to begin"). Tom authored early logs only; his LAST daily log 2025-04-08 (tree/palm selections) - Tom exits construction ~1 month before the 2025-05-09 termination; Scott Culver dominant PM thereafter. NO work stoppage around 2025-05-09 termination or 2025-05-19 warning meeting - logs dense/continuous (excavation done ~May 3-5, plumbing May 8, steel May 12-16, mister trenching May 19, gunite May 28). 2026 logs = warranty/service tickets repeating "Start up date: 11/10/25". Feeds EVT-0031..0037.'}
for row,d in [(22,d119),(23,d120)]:
    for k,v in d.items(): P(row,k,v)
    print('added',d['Source ID'],'row',row)
wb.save(WB); print('SAVED')
