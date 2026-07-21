import sys, os, shutil
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map, set_literal
from openpyxl import load_workbook

M = r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'
WB = os.path.join(M, 'Wilson.xlsx'); SC = os.path.join(M, 'Wilson.schema.json')
assert not os.path.exists(os.path.join(M, '~$Wilson.xlsx')), "OPEN in Excel"
bk = os.path.join(M, 'Wilson_prewrite_backup_2026-07-20_CD19_MEETING.xlsx')
shutil.copy2(WB, bk); print("BACKUP", bk)
schema = load_json_strict(SC); wb = load_workbook(WB)
cl = wb['CLAIMS & DEFENSES']; cc = col_map(schema, 'CLAIMS & DEFENSES')
assert cl.cell(row=24, column=cc['ID']).value == 'CD-0019', "CD-0019 not at row 24"
note = ("THEORY per AB account 2026-07-20; counsel to refine legal framing and elements. "
 "WARNING MEETING IDENTIFIED: Michael Baker + Scott Culver met David Wilson in person at Wilson's house Monday 2025-05-19, 9:30-10:00 AM, "
 "to warn him off Tom Delaney (the 'cornering' Wilson later referenced); originally scheduled earlier, pushed from ~2025-05-16 because "
 "Wilson was 'in bed sick' (Managers chat, Michael Baker) - per Gemini reading of internal chats/transcripts; transcript + Managers chat to be preserved. "
 "Evidence located/collected: (a) SRC-0103 Wilson 2025-09-11 'Tom stole over $95k from me... thank you for the help'; "
 "(b) SRC-0109 CO grid incl CO-0036 Ramada handoff to Bradley Schneider; (c) Wilson 2025-05-24 'Re: Ramada Change' emails - Wilson coordinating Schneider himself; "
 "(d) Wilson 2025-07-29 email to Wayne Parker / JE Pierre LLC re pool square footage; (e) Wilson 2026-01-26 emails to Heath Richards @roc.az.gov "
 "(ROC complaint 2026-00144, re pool interior); (f) Wilson Las Vegas address + 2025 emails in +0900/+0800 (Asia) timezones = absentee/overseas owner; "
 "(g) SRC-0115 Tom Cease and Desist; (h) Veronica 2025-05-16 vendor revocation notice. "
 "TO COLLECT/PRESERVE: Read AI/Gemini transcript of the 2025-05-19 meeting + Managers chat 2025-05-16; Wilson's 'why are you doing this to Tommy' message "
 "(likely text/voice to Michael); garage-overlay + auction-bidding specifics.")
set_literal(cl.cell(row=24, column=cc['Notes']), note)
print("CD-0019 note updated with 2025-05-19 warning meeting + timeline")
wb.save(WB); print("SAVED")
