import sys, os, shutil, hashlib
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map, set_literal
from openpyxl import load_workbook

M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'
WB=M+r'\Wilson.xlsx'; SC=M+r'\Wilson.schema.json'
KEY=M+r'\preserved\keydocs\SRC-0118_2026-07-19_Wilson-cornered.eml'
assert not os.path.exists(M+r'\~$Wilson.xlsx'), "OPEN in Excel"
sha=hashlib.sha256(open(KEY,'rb').read()).hexdigest(); print('SRC-0118 SHA=',sha)
bk=M+r'\Wilson_prewrite_backup_2026-07-20_SRC0118.xlsx'; shutil.copy2(WB,bk); print('BACKUP',bk)
s=load_json_strict(SC); wb=load_workbook(WB); si=wb['SOURCE INDEX']; c=col_map(s,'SOURCE INDEX')
r=None
for rr in range(6, si.max_row+2):
    if not si.cell(row=rr,column=c['Source ID']).value: r=rr; break
assert r==19, 'expected next row 19, got %s'%r
def put(col,val): set_literal(si.cell(row=r,column=c[col]), val)
note=('KEY ADMISSION. Wilson (davidwilson0@yahoo.com) 2026-07-19 8:51 AM to Michael Baker; Michael fwd to Aaron/Martin/Christine 2026-07-19 16:30. '
 'Verbatim: "The fact you, Scott and Martin came to my house and cornered me regarding Tom\'s termination and criminal behavior was not necessary... '
 'How many times does a company send their lead council to a client\'s house to discuss the termination of an employee???" '
 'Wilson himself places the 2025-05-19 warning meeting and concedes Omni warned him re Tom\'s termination + criminal behavior. '
 'Also attributes defects to "Tom\'s criminal and deviant behavior... which I have corrected without Omni taking any responsibility" '
 '(pebble-shine finish, glass tiles, rubber mulch, misters, fiber optic, cave door); threatens to hire another contractor and warns of "severe damage" (escalation posture). '
 'Supports CD-0016 (notice) and CD-0019 (warned + engaged Tom anyway). NATIVE .eml + canonical hash PENDING forensic re-export. '
 'Same-thread items to preserve: Michael 2026-07-17 letter (attachment "David Wilson.pdf" = Omni position); Michael 2026-07-20 reply (final payment released via BT 2026-03-16).')
data={'Source ID':'SRC-0118','SHA-256':sha,'Doc Date':'2026-07-19','Date Type':'SENT',
 'Date Basis':'RFC 5322 Date header (sender-supplied; unverified) - connector capture, native RFC822 pending',
 'Type':'Email','Description':'Wilson 2026-07-19 "cornered" email (fwd by Michael to ownership) - Wilson concedes Omni came to his house re Tom termination',
 'Custodian':'Aaron Baker (aaron mailbox); fwd by Michael Baker','From / Author':'David Wilson <davidwilson0@yahoo.com> (fwd: Michael Baker <michael@omnipools.com>)',
 'To / Recipients':'Michael Baker <michael@omnipools.com>; fwd to aaron@omnipools.com, martin@, christine.stewart@',
 'Native (relative URI)':'gmail:thread/19f72526564ff643 msg/19f7b36f8054112b','Working Copy (filename)':'SRC-0118_2026-07-19_Wilson-cornered.eml',
 'Acquisition Method':'Located + retrieved via Gmail API connector 2026-07-20 (Aaron mailbox). Content-faithful capture; message postdates the 2026-07-17 forensic export. Native RFC822 + canonical SHA pending forensic re-export.',
 'Received Date':'2026-07-19','Disposition':'RECEIVED','Notes':note}
for k,v in data.items(): put(k,v)
print('SRC-0118 added at row',r)
# CD-0019 confirm
cl=wb['CLAIMS & DEFENSES']; cc=col_map(s,'CLAIMS & DEFENSES'); nc=cc['Notes']
assert cl.cell(row=24,column=cc['ID']).value=='CD-0019'
addc=(' CONFIRMED 2026-07-20 via Gmail (SRC-0118): Wilson\'s 2026-07-19 email states verbatim "you, Scott and Martin came to my house and cornered me '
 'regarding Tom\'s termination and criminal behavior." Wilson himself places the warning meeting and admits Omni warned him re Tom. '
 'NOTE attendee variance: Wilson recalls Michael+Scott+Martin; internal records (Gemini 49/73) show Michael+Scott - reconcile.')
set_literal(cl.cell(row=24,column=nc),(str(cl.cell(row=24,column=nc).value or '').rstrip()+addc).strip())
print('CD-0019 confirmed')
wb.save(WB); print('SAVED')
