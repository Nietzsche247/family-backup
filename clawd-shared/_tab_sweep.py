import sys, re
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from openpyxl import load_workbook
WB=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1\Wilson.xlsx'
wb=load_workbook(WB)
idpat=re.compile(r'^[A-Z]{2,5}-\d{3,4}$')
print('%-20s %7s %5s %6s %6s  %s'%('SHEET','dims','pop','formu','data','id-records'))
for ws in wb.worksheets:
    f=d=pop=0; ids=[]
    for row in ws.iter_rows():
        has=False
        for cell in row:
            v=cell.value
            if v is None: continue
            has=True
            if isinstance(v,str) and v.startswith('='): f+=1
            else: d+=1
        if has: pop+=1
    for r in range(1, ws.max_row+1):
        v=ws.cell(row=r,column=1).value
        if v is not None and idpat.match(str(v).strip()): ids.append(str(v).strip())
    dims='%dx%d'%(ws.max_row,ws.max_column)
    rng=(ids[0]+'..'+ids[-1]) if ids else '-'
    print('%-20s %7s %5d %6d %6d  n=%d %s'%(ws.title[:20],dims,pop,f,d,len(ids),rng))
