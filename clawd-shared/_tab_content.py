import sys
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from openpyxl import load_workbook
WB=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1\Wilson.xlsx'
wb=load_workbook(WB)
targets=['KEY DATES','FACT REGISTRY','THIRD PARTIES','DAMAGES','DEADLINES','LITIGATION LOG','COST LEDGER','DISCLOSURE REVIEW','STAGING','IMPORT LANDING','AUDIT','ORGANIZATIONS','CAST','DISPUTED FACTS']
out=[]
def cellval(v):
    if v is None: return ''
    s=str(v)
    if s.startswith('='): return '{f}'
    return s[:22]
for t in targets:
    ws=wb[t]
    out.append('\n===== %s  (%dx%d) ====='%(t,ws.max_row,ws.max_column))
    shown=0
    for r in range(1, ws.max_row+1):
        vals=[ws.cell(row=r,column=col).value for col in range(1, min(ws.max_column,6)+1)]
        nonform=[v for v in vals if v is not None and not (isinstance(v,str) and v.startswith('='))]
        # show first 6 rows (structure) OR any row with real (non-formula) data in col1/col2
        c1nf = vals[0] is not None and not (isinstance(vals[0],str) and str(vals[0]).startswith('='))
        c2nf = len(vals)>1 and vals[1] is not None and not (isinstance(vals[1],str) and str(vals[1]).startswith('='))
        if r<=6 or c1nf or c2nf:
            if shown<34:
                out.append('  r%-3d | %s'%(r,' | '.join(cellval(v) for v in vals)))
                shown+=1
    if shown>=34: out.append('   ... (truncated)')
open(r'C:\Users\aaron\clawd-shared\_tab_content.txt','w',encoding='utf-8').write('\n'.join(out))
print('done, sheets dumped:', len(targets))
