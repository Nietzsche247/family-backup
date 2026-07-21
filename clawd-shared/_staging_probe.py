import sys
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map
from openpyxl import load_workbook
M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'
s=load_json_strict(M+r'\Wilson.schema.json'); wb=load_workbook(M+r'\Wilson.xlsx')
ws=wb['STAGING']; cm=col_map(s,'STAGING')
print('STAGING COLS:', cm)
print('dims %dx%d'%(ws.max_row,ws.max_column))
print('--- rows 1-8 (structure) ---')
for r in range(1,9):
    vals=[ws.cell(row=r,column=col).value for col in range(1,min(ws.max_column,len(cm))+1)]
    print(' r%d'%r, [ (str(v)[:20] if v is not None else '') for v in vals])
# find first empty data row (col1 blank at/after row 6)
nxt=None
for r in range(6, ws.max_row+2):
    if not ws.cell(row=r,column=1).value: nxt=r; break
print('next empty row =', nxt)
print('staging_status enum =', s['enums'].get('staging_status'))
# check if MASTER CHRONOLOGY requires Source ID (look at audit-relevant): show a dateless event row's Source to confirm sourced
