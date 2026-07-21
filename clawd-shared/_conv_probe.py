import sys, os
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map
from openpyxl import load_workbook
M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'
s=load_json_strict(M+r'\Wilson.schema.json'); wb=load_workbook(M+r'\Wilson.xlsx')
ws=wb['SOURCE INDEX']; cm=col_map(s,'SOURCE INDEX')
def full(r,k): 
    v=ws.cell(row=r,column=cm[k]).value; return v
for sid,r in [('SRC-0101',7),('SRC-0105',11),('SRC-0115',16)]:
    print('===',sid,'(row',r,')')
    for k in ['Doc Date','Date Type','Date Basis','Native (relative URI)','Working Copy (filename)','Acquisition Method','Description','Notes']:
        print('   %-24s: %r'%(k, full(r,k)))
print()
print('=== preserved/ tree ===')
base=M+r'\preserved'
for root,dirs,files in os.walk(base):
    lvl=root.replace(base,'').count(os.sep)
    print('  '*lvl, os.path.basename(root)+'/')
    for f in files: print('  '*(lvl+1), f)
