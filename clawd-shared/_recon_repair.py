import sys, os, json, shutil
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map, set_literal
from openpyxl import load_workbook
M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'; SRC=M+r'\Wilson.xlsx'; COPY=M+r'\Wilson_repaired.xlsx'
shutil.copy2(SRC,COPY); print('working copy ->',COPY)
recs=json.load(open(r'C:\Users\aaron\clawd-shared\_recon_data.json',encoding='utf-8'))
assert len(recs)==12 and recs[0]['A']=='R-0001' and recs[11]['A']=='R-0012', 'record set wrong'
s=load_json_strict(M+r'\Wilson.schema.json'); wb=load_workbook(COPY)
ws=wb['RECONCILIATION']
HEADERS=['Line ID','Category','Item','Their Transaction ID','Their $ (auto)','Approval Event ID','Approval Documented? (auto)','Payment Transaction ID','Payment Documented? (auto)','Our Transaction ID','Our $ (auto)','Delta (auto)','Status (auto)','Counsel Status','Notes']
HUMAN={'A':1,'B':2,'C':3,'D':4,'F':6,'H':8,'J':10,'N':14,'O':15}
FORM={5:'=IF($D{r}="","",IFERROR(INDEX(TRANSACTIONS!$D$6:$D$605,MATCH($D{r},TRANSACTIONS!$A$6:$A$605,0)),""))',
 7:'=IF($C{r}="","",IF($F{r}="","N","Y"))',
 9:'=IF($C{r}="","",IF($H{r}="","N","Y"))',
 11:'=IF($J{r}="","",IFERROR(INDEX(TRANSACTIONS!$D$6:$D$605,MATCH($J{r},TRANSACTIONS!$A$6:$A$605,0)),""))',
 12:'=IF(OR(NOT(ISNUMBER($E{r})),NOT(ISNUMBER($K{r}))),"",$K{r}-$E{r})',
 13:'=IF($C{r}="","",IF(AND($G{r}="Y",$I{r}="Y"),IF(AND(ISNUMBER($E{r}),ISNUMBER($K{r}),$E{r}<>$K{r}),"APPROVAL AND PAYMENT DOCUMENTED; AMOUNT DIFFERS","APPROVAL AND PAYMENT DOCUMENTED"),IF(AND(ISNUMBER($E{r}),ISNUMBER($K{r})),IF($E{r}=$K{r},"AMOUNTS MATCH","AMOUNT DIFFERS"),IF($K{r}="","NO RECORD YET","COUNSEL REVIEW"))))'}
# 1) clear rows 6,7 fully; clear row 8 and write header
for r in (6,7,8):
    for col in range(1,16): ws.cell(row=r,column=col).value=None
for col,name in enumerate(HEADERS,1): set_literal(ws.cell(row=8,column=col),name)
# 2) records into rows 9-20 (human verbatim + schema formulas)
for i,rec in enumerate(recs):
    r=9+i
    for lab,col in HUMAN.items():
        v=rec.get(lab,'')
        ws.cell(row=r,column=col).value = None if (v=='' or v is None) else None
        if not (v=='' or v is None): set_literal(ws.cell(row=r,column=col),v)
    for col,tmpl in FORM.items(): ws.cell(row=r,column=col).value=tmpl.replace('{r}',str(r))
print('placed R-0001..R-0012 at rows 9-20; header at row 8; rows 6-7 cleared')
# 3) table columns must match header
tbl=ws.tables['tblRecon']; print('tblRecon ref',tbl.ref,'cols',len(tbl.tableColumns))
assert len(tbl.tableColumns)==15
for i,name in enumerate(HEADERS): tbl.tableColumns[i].name=name
print('tblRecon column names synced to header')
# 4) CD link review (append only; do NOT touch Status/Counsel Status)
cd=wb['CLAIMS & DEFENSES']; cc=col_map(s,'CLAIMS & DEFENSES')
def rowof(idv):
    for r in range(1,cd.max_row+1):
        if cd.cell(row=r,column=1).value==idv: return r
def app_note(r,extra):
    cur=cd.cell(row=r,column=cc['Notes']).value or ''
    set_literal(cd.cell(row=r,column=cc['Notes']),(str(cur).rstrip()+extra).strip())
r15=rowof('CD-0015')
set_literal(cd.cell(row=r15,column=cc['Supporting Event IDs']),'EVT-0014; EVT-0016; EVT-0041')
app_note(r15,' | EARLIER NOTICE: EVT-0041 (2025-05-19 in-person warning at Wilson home) is an earlier notice-of-termination record - contemporaneous chat SRC-0131 + Wilson admission SRC-0118. (Status/Counsel Status unchanged for counsel.)')
r18=rowof('CD-0018')
set_literal(cd.cell(row=r18,column=cc['Supporting Event IDs']),'EVT-0022; EVT-0023; EVT-0024; EVT-0025')
app_note(r18,' | LINKED 2026-07-21: scope-removal events EVT-0022 (CO-0028), EVT-0023 (CO-0036 Ramada), EVT-0024 (CO-0048 landscaping), EVT-0025 (CO-0051); transactions TXN-0013..0016. (Status/Counsel Status unchanged for counsel.)')
r19=rowof('CD-0019')
set_literal(cd.cell(row=r19,column=cc['Supporting Event IDs']),'EVT-0041')
app_note(r19,' | LINK/FLAG 2026-07-21: SUPPORTED - EVT-0041 (2025-05-19 warning meeting; SRC-0131 contemporaneous chat + SRC-0118 Wilson admission) supports the express-warning assertion. UNLINKED (no chronology event yet, flagged for counsel): "banned Delaney from the job repeatedly"; the written cease-and-desist (preserved as SRC-0115 but no EVT); Wilson\'s independent engagement of Delaney (partial support SRC-0103 + the CO scope-removal events). (Status/Counsel Status unchanged.)')
print('CD-0015/0018/0019 links added (Status/Counsel Status untouched)')
wb.save(COPY); print('SAVED copy')
