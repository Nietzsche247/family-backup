import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'
P=M+r'\Wilson_Counsel_Strategy_v3.xlsx'
wb=load_workbook(P)
print('SHEETS:', wb.sheetnames)
for name in wb.sheetnames:
    ws=wb[name]
    print('\n' + '='*70)
    print('SHEET: %s   (%d rows x %d cols)'%(name, ws.max_row, ws.max_column))
    print('='*70)
    for r in range(1, min(ws.max_row,200)+1):
        cells=[]
        for col in range(1, min(ws.max_column,15)+1):
            v=ws.cell(row=r,column=col).value
            if v in (None,''): continue
            pref='=' if (isinstance(v,str) and v.startswith('=')) else ''
            cells.append('%s:%s%s'%(get_column_letter(col),pref,str(v)[:150]))
        if cells: print(' r%-3d %s'%(r,' | '.join(cells)))
