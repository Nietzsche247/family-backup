from openpyxl import load_workbook
M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'
P=M+r'\Wilson_Counsel_Strategy_FINAL_2026-07-21.xlsx'
wb=load_workbook(P)
STALE=['structural repair required','RECONCILIATION structural defect','Repair RECONCILIATION',
       'lack supporting Event IDs even though','Run the promised native','it missed the malformed',
       'Several newer events are not yet linked','Replace it with this generated v3',
       'returned 15 SE-WB-03','Wilson.schema.json, MATTER_INSTANCE.json']
hits=[]
for sh in wb.sheetnames:
    ws=wb[sh]
    for row in ws.iter_rows():
        for cell in row:
            v=cell.value
            if isinstance(v,str):
                for s in STALE:
                    if s.lower() in v.lower(): hits.append((sh,cell.coordinate,s))
print('STALE-TEXT SCAN:', 'CLEAN - no stale statements found' if not hits else hits)
# preserved limitations must still be present
KEEP=['FACT REGISTRY','DISPUTED FACTS','DEADLINES has 0 rows','DRAFT','targeted, not exhaustive',
      '2026-07-22','1.39 GB','4/27','15,375.58']
found={k:False for k in KEEP}
for sh in wb.sheetnames:
    for row in wb[sh].iter_rows():
        for cell in row:
            if isinstance(cell.value,str):
                for k in KEEP:
                    if k.lower() in cell.value.lower(): found[k]=True
print('\nHONEST-LIMITATION RETENTION CHECK:')
for k,v in found.items(): print('   %-24s %s'%(k,'PRESENT' if v else '*** MISSING ***'))
print('\nkey cells:')
print(' EXEC A4  :', wb['EXECUTIVE']['A4'].value)
tr=wb['TECHNICAL READINESS']
for r in (8,9,10,20):
    print(' TR r%-2d   : %s | %s'%(r, tr['B%d'%r].value, tr['C%d'%r].value))
print(' TR r18   : %s | %s'%(tr['B18'].value, tr['C18'].value))
cm=wb['CLAIM MAP']
for r in (20,23,24):
    print(' CM r%-2d   : %s  SUP=%s  STATUS=%s  COUNSEL=%s'%(r, cm['A%d'%r].value, cm['E%d'%r].value, cm['G%d'%r].value, cm['H%d'%r].value))
