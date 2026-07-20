import sys, os, shutil
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map, set_literal
from openpyxl import load_workbook

M = r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'
WB = os.path.join(M, 'Wilson.xlsx'); SC = os.path.join(M, 'Wilson.schema.json')
schema = load_json_strict(SC); wb = load_workbook(WB)
si = wb['SOURCE INDEX']; sc = col_map(schema, 'SOURCE INDEX'); idc = sc['Source ID']

# 1) clear the misplaced SRC-0115 at row 4 (all 20 cols)
assert si.cell(row=4, column=idc).value == 'SRC-0115', "row4 not SRC-0115"
for c in range(1, 21):
    si.cell(row=4, column=c).value = None
print("cleared misplaced row 4")

# 2) real next empty DATA row (header at row 5, data from row 6)
r = 6
while si.cell(row=r, column=idc).value not in (None, ''): r += 1
assert r >= 16, f"unexpected next row {r}"
print("correct next data row:", r)

def putcol(names, value):
    for nm in names:
        if nm in sc:
            set_literal(si.cell(row=r, column=sc[nm]), value); return
    raise KeyError(names)
putcol(['Source ID'], 'SRC-0115')
putcol(['Type'], 'Correspondence')
putcol(['Description'], 'Cease-and-desist letter issued by Omni to Tom Delaney (file: Tom Cease and Desist.pdf)')
putcol(['Custodian'], 'Christine Stewart')
putcol(['From/Author', 'From / Author'], 'Omni Pool Builders & Design LLC')
putcol(['To/Recipients', 'To / Recipients'], 'Tom Delaney')
putcol(['Disposition'], 'RECEIVED')
putcol(['Notes'], ("SLOT for Christine Stewart to attach the canonical/signed Tom Cease and Desist.pdf and set SHA-256, Working Copy, "
    "Doc Date, and final Disposition. A copy is already located in the preserved email corpus as an attachment to "
    "Christine->Michael emails: 'Tom Delaney Documents' 2025-07-14 (christine.stewart/1980b0c01d3baa20.eml) and "
    "'Tom Delaney - Write Ups, Termination etc' 2026-07-13 (christine.stewart/19f5cbc7b9d1129e.eml). Christine to confirm "
    "this is the final version or supply the signed original. Supports CD-0017 and CD-0019. SRC-0110..0114 reserved for pending invoice/payment exports."))
print("SRC-0115 rewritten at row", r)
wb.save(WB); print("SAVED")
