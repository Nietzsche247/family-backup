import sys
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map
from openpyxl import load_workbook
M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'
s=load_json_strict(M+r'\Wilson.schema.json'); wb=load_workbook(M+r'\Wilson.xlsx')
ws=wb['MASTER CHRONOLOGY']
print('ALL ENUM KEYS:', list(s.get('enums',{}).keys()))
for k in s.get('enums',{}):
    kl=k.lower()
    if any(x in kl for x in ['date','tag','statement','issue','auth','admiss','corrob','transc','location']):
        print('  ENUM',k,'=',s['enums'][k])
print()
print('CHECK formula (r24 c30):')
print('  ', ws.cell(row=24,column=30).value)
print('HOTRANK formula (r24 c31):')
print('  ', ws.cell(row=24,column=31).value)
