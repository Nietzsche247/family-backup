# read-only deep probe: COVERAGE, CAST, TRANSACTIONS, RECONCILIATION win-meter, cached-value test
import openpyxl
P = r"C:\North_Star_Projects\WILSON_MATTER_WORKING_v1\Wilson.xlsx"

def dump(ws, label, maxr=60, maxc=16):
    print(f"\n===== {label} =====")
    for r in ws.iter_rows(min_row=1, max_row=maxr, max_col=maxc):
        cells = []
        for c in r:
            v = c.value
            if v is None: 
                continue
            s = str(v)
            if len(s) > 40: s = s[:40]+"..."
            cells.append(f"{c.coordinate}={s}")
        if cells:
            print(" | ".join(cells))

wb = openpyxl.load_workbook(P, data_only=False, read_only=True)
for name in ["COVERAGE","CAST","TRANSACTIONS"]:
    if name in wb.sheetnames:
        dump(wb[name], name)
wb.close()

# cached-value test: are formulas computed or inert?
wb2 = openpyxl.load_workbook(P, data_only=True, read_only=True)
print("\n===== CACHED-VALUE TEST (data_only=True) =====")
for name in ["START","MASTER CHRONOLOGY","RECONCILIATION","DAMAGES","COVERAGE"]:
    if name not in wb2.sheetnames: 
        continue
    ws = wb2[name]
    total=0; cached=0
    for r in ws.iter_rows():
        for c in r:
            # only look where the formula wb had a formula
            pass
    # simpler: count non-None numeric/text in first 40 rows
    nn=0
    for r in ws.iter_rows(min_row=1,max_row=40,max_col=16):
        for c in r:
            if c.value is not None: nn+=1
    print(f"{name:<22} non-None cells (rows1-40) when data_only=True: {nn}")
wb2.close()
