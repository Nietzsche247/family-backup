import sys, os, json
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'
wb=load_workbook(M+r'\Wilson.xlsx')  # formulas visible
ws=wb['RECONCILIATION']
print('=== TABLES on RECONCILIATION ===')
for t in ws.tables.values(): print('  %s  ref=%s'%(t.name, t.ref))
print('dims: %d rows x %d cols'%(ws.max_row, ws.max_column))
print('\n=== ROWS 1-24, cols A-O (F=formula) ===')
for r in range(1,25):
    cells=[]
    for col in range(1,16):
        v=ws.cell(row=r,column=col).value
        if v in (None,''): continue
        pref='=' if (isinstance(v,str) and v.startswith('=')) else ''
        sv=str(v)
        cells.append('%s:%s%s'%(get_column_letter(col), pref, sv[:32]))
    if cells: print(' r%-3d %s'%(r,' | '.join(cells)))
print('\n=== locate R-0001..R-0012 in col A ===')
for r in range(1,ws.max_row+1):
    a=ws.cell(row=r,column=1).value
    if isinstance(a,str) and a.startswith('R-0'): print('  %s at row %d'%(a,r))
# schema defs
print('\n=== SCHEMA: RECONCILIATION columns ===')
s=json.load(open(M+r'\Wilson.schema.json',encoding='utf-8'))
sh=s.get('sheets',s)
node=sh.get('RECONCILIATION')
if node:
    for i,cd in enumerate(node.get('columns',[]),1):
        nm=cd.get('name'); f=cd.get('formula') or cd.get('auto_formula') or cd.get('expr')
        extra='  FORMULA='+str(f) if f else ('  [auto]' if cd.get('auto') or (nm and 'auto' in nm.lower()) else '')
        print('  col%2d %-26s%s'%(i,nm,extra))
    print('  header_row?', node.get('header_row'), ' table?', node.get('table'), node.get('table_ref'))
print('\n=== scripts present ===')
sd=r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts'
for f in sorted(os.listdir(sd)):
    if f.endswith('.py'): print('  ',f)
