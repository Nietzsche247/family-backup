from openpyxl import load_workbook
M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'
wb=load_workbook(M+r'\Wilson_Counsel_Strategy_v3.xlsx')
T={'EXECUTIVE':[('A',4),('B',7),('B',8),('D',19),('D',22)],
   'TECHNICAL READINESS':[('A',5),('C',8),('D',8),('E',8),('C',9),('D',9),('E',9),('C',10),('D',10),('E',10),('C',12),('D',12),('E',12),('C',18),('D',18),('E',18)],
   'OPEN PROOF':[('B',19),('B',22),('B',23)],
   'CLAIM MAP':[('I',20),('I',23),('I',24)]}
for sh,cells in T.items():
    ws=wb[sh]; print('\n########## %s ##########'%sh)
    for col,r in cells:
        v=ws['%s%d'%(col,r)].value
        print('\n[%s%d] %s'%(col,r,v))
