import sys, os, shutil
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map, set_literal
from openpyxl import load_workbook
M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'; WB=M+r'\Wilson.xlsx'; SC=M+r'\Wilson.schema.json'
assert not os.path.exists(M+r'\~$Wilson.xlsx'), "OPEN in Excel"
bk=M+r'\Wilson_prewrite_backup_2026-07-21_STAGING.xlsx'; shutil.copy2(WB,bk); print('BACKUP',bk)
s=load_json_strict(SC); wb=load_workbook(WB); ws=wb['STAGING']; c=col_map(s,'STAGING')
rows=[
 {'What we believe happened':'Omni sent David Wilson an EMAIL requesting the final payment (5% retention / invoice 0007, $51,752.62).',
  'Memory Source':'Michael Baker; corroborated by Michael 2026-07-20 email ("final payment was requested a few times")',
  'What Document Proves It':'The sent final-payment-request email; Michael 2026-07-20 reply references it',
  'Where to Look':'Gmail Sent (michael@/aaron@); Buildertrend messages + invoice 0007 send log','Status':'OPEN'},
 {'What we believe happened':'Omni TEXTED David Wilson a reminder that the final payment (5% retention) was due (text #1).',
  'Memory Source':'Michael Baker','What Document Proves It':'Text thread with Wilson (714-269-6161) - screenshot/export',
  'Where to Look':"Michael's phone; text thread with 714-269-6161",'Status':'OPEN'},
 {'What we believe happened':'Omni sent a SECOND text reminder to David Wilson about the outstanding final payment (text #2).',
  'Memory Source':'Michael Baker','What Document Proves It':'Text thread with Wilson (714-269-6161) - screenshot/export',
  'Where to Look':"Michael's phone; second text in the thread with 714-269-6161",'Status':'OPEN'},
 {'What we believe happened':'Michael Baker emailed David Wilson stating the warranties will be RE-ESTABLISHED once the final payment has been made.',
  'Memory Source':'Michael Baker; Aaron','What Document Proves It':'The sent email conditioning warranty reinstatement on final payment',
  'Where to Look':"Gmail Sent (michael@); 'Warranties and job completion items' thread; possibly the 2026-07-17 letter (David Wilson.pdf) or 2026-07-20 reply",'Status':'OPEN'},
]
start=6
for i,d in enumerate(rows):
    r=start+i
    assert not ws.cell(row=r,column=1).value, 'row %d not empty'%r
    for k,v in d.items(): set_literal(ws.cell(row=r,column=c[k]),v)
    print('staged row',r,'-',d['What we believe happened'][:50])
wb.save(WB); print('SAVED')
