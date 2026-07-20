import sys, os
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map
from openpyxl import load_workbook

WB = r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1\Wilson.xlsx'
SC = r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1\Wilson.schema.json'

# is the live workbook open in Excel?
lock = os.path.join(os.path.dirname(WB), '~$'+os.path.basename(WB))
print("WORKBOOK LOCK PRESENT:", os.path.exists(lock))

schema = load_json_strict(SC)
for s in ['SOURCE INDEX','RECONCILIATION','TRANSACTIONS']:
    try:
        print(f"\n--- col_map {s} ---")
        print(col_map(schema, s))
    except Exception as e:
        print(s, "ERR", repr(e))

wb = load_workbook(WB, data_only=False, read_only=True)

print("\n===== SOURCE INDEX (rows 5-22, non-formula) =====")
si = wb['SOURCE INDEX']
for r in si.iter_rows(min_row=5, max_row=22):
    vals = [(c.coordinate, str(c.value)[:45]) for c in r if c.value is not None and not (isinstance(c.value,str) and c.value.startswith('='))]
    if vals: print(vals)

print("\n===== RECONCILIATION (rows 5-20, non-formula) =====")
rec = wb['RECONCILIATION']
for r in rec.iter_rows(min_row=5, max_row=20):
    vals = [(c.coordinate, str(c.value)[:60]) for c in r if c.value is not None and not (isinstance(c.value,str) and c.value.startswith('='))]
    if vals: print(vals)
wb.close()

# schema enums for SOURCE INDEX + RECONCILIATION if present
print("\n===== schema keys =====")
print(type(schema), list(schema.keys())[:10] if isinstance(schema, dict) else 'notdict')
