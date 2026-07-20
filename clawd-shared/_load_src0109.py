import sys, os, shutil, datetime, hashlib
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map, set_literal
from openpyxl import load_workbook

M = r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'
WB = os.path.join(M, 'Wilson.xlsx'); SC = os.path.join(M, 'Wilson.schema.json')
assert not os.path.exists(os.path.join(M, '~$Wilson.xlsx')), "OPEN in Excel"

bk = os.path.join(M, 'Wilson_prewrite_backup_2026-07-20_SRC0109.xlsx')
shutil.copy2(WB, bk); print("BACKUP", bk)

src = r'C:\Users\aaron\Downloads\ChangeOrders.xlsx'
stg = os.path.join(M, 'staging4_financial', 'ChangeOrders_WILSON_grid_2026-07-20.xlsx')
shutil.copy2(src, stg)
os.makedirs(os.path.join(M, 'preserved', 'financial'), exist_ok=True)
wc = os.path.join(M, 'preserved', 'financial', 'SRC-0109_change-order-grid_WILSON.xlsx')
shutil.copy2(src, wc)
sha = hashlib.sha256(open(wc, 'rb').read()).hexdigest()
mt = datetime.datetime.fromtimestamp(os.path.getmtime(src), tz=datetime.timezone.utc).isoformat()
print("SRC-0109 sha", sha)

schema = load_json_strict(SC); wb = load_workbook(WB)

si = wb['SOURCE INDEX']; sc = col_map(schema, 'SOURCE INDEX'); R = 15
assert si.cell(row=R, column=sc['Source ID']).value in (None, ''), "SI row15 not empty"
sivals = {'Source ID':'SRC-0109','SHA-256':sha,'Doc Date':'2026-07-20','Date Type':'CAPTURE',
 'Date Basis':"Export generated 2026-07-20 (grid header 'Change Orders (exported on Mon, Jul 20, 2026)').",
 'Type':'Change Order',
 'Description':'Buildertrend Change Orders grid export (Excel), Wilson job 38668322. 56 change orders CO-0001..CO-0062 with title, amount, approval status and date, created-by, and client. Grid subtotal $538,179.12; builder cost $338,450.85.',
 'Custodian':'aaron@omnipoolbuilders.com','From / Author':'Buildertrend (Omni builder acct 6769), system export',
 'Native (relative URI)':'dir:staging4_financial/ChangeOrders_WILSON_grid_2026-07-20.xlsx',
 'Working Copy (filename)':'SRC-0109_change-order-grid_WILSON.xlsx',
 'Acquisition Method':'A1 administrative export; Buildertrend Change Orders Export all -> Export to Excel (POST /api/ChangeOrders/ExportToExcel 200); Wilson job 38668322; exported 2026-07-20.',
 'Received Date':'2026-07-20','Disposition':'PRESERVED',
 'Notes':'filesystem mtime '+mt+'. 56 CO rows confirmed in code. Four scope deducts CO-0036 -$63,917.18, CO-0048 -$114,857.80, CO-0028 -$3,061.27, CO-0051 -$1,279.81 (combined -$183,116.06); each Status "Approved on <date>", Client "David Wilson". Export self-identifies the job: Client column = David Wilson on every row.'}
for f, v in sivals.items():
    if f in sc: set_literal(si.cell(row=R, column=sc[f]), v)
print("SRC-0109 written row 15")

tx = wb['TRANSACTIONS']; tc = col_map(schema, 'TRANSACTIONS')
deducts = [
 ('TXN-0013','2025-05-24','-63917.18','CO-0036 "Schneider to own the entire Ramada and finishes". Client-approved scope removal from Omni contract. Created by Michael Baker 2025-05-23.','Approved by client David Wilson on 2025-05-24 (created 2025-05-23) per Buildertrend CO grid export SRC-0109. Grid Status: "Approved on 5-24-2025". Approval postdates the pleaded termination date 2025-05-09.'),
 ('TXN-0014','2025-08-11','-114857.80','CO-0048 "Remove Landscaping packages from Omni Scope". Client-approved scope removal. Created 2025-08-11.','Approved by client David Wilson on 2025-08-11 per SRC-0109. Grid Status: "Approved on 8-11-2025".'),
 ('TXN-0015','2025-04-03','-3061.27','CO-0028 "Eq, Lights, Water Softner, and pumps - POND". Client-approved deduct (client total -$3,061.27; builder cost -$7,134.11). Created 2025-03-28.','Approved by client David Wilson on 2025-04-03 per SRC-0109. Grid Status: "Approved on 4-3-2025".'),
 ('TXN-0016','2025-08-29','-1279.81','CO-0051 "Remove gate from Scope and Flipping one gate". Client-approved deduct. Created 2025-08-27.','Approved by client David Wilson on 2025-08-29 per SRC-0109. Grid Status: "Approved on 8-29-2025".'),
]
row = 18
for tid, dt, amt, desc, note in deducts:
    assert tx.cell(row=row, column=tc['Transaction ID']).value in (None, ''), f"TX row {row} not empty"
    tvals = {'Transaction ID':tid,'Date':dt,'Type':'CHANGE','Amount':amt,'Currency':'USD',
     'From Actor ID':'ACT-0001','From Org ID':'ORG-0001','Description':desc,
     'Source ID(s)':'SRC-0109','Status':'SOURCE-MATCHED','Notes':note}
    for f, v in tvals.items():
        if f in tc: set_literal(tx.cell(row=row, column=tc[f]), v)
    row += 1
print("TXN-0013..0016 written rows 18-21")

rec = wb['RECONCILIATION']; rc = col_map(schema, 'RECONCILIATION')
def setrec(rn, eid, otxn, note):
    assert rec.cell(row=rn, column=rc['Line ID']).value == eid, f"{eid} not at row {rn}"
    if otxn: set_literal(rec.cell(row=rn, column=rc['Our Transaction ID']), otxn)
    set_literal(rec.cell(row=rn, column=rc['Notes']), note)
setrec(8,'R-0003',None,'COLLECTED 2026-07-20 as SRC-0109 (Buildertrend CO grid export, 56 records). Grid Subtotal $538,179.12; Total price $538,179.12; Builder cost $338,450.85. All rows Client = David Wilson.')
setrec(9,'R-0004','TXN-0013','COLLECTED as SRC-0109; TXN-0013. CO-0036 -$63,917.18 "Schneider to own the entire Ramada and finishes". Client-approved by David Wilson 2025-05-24 (created 2025-05-23), 15 days after the pleaded termination 2025-05-09. Prior note said 2025-05-23; that is the created date - approval is 2025-05-24 per grid Status. Counsel Status HOLD pending review.')
setrec(10,'R-0005','TXN-0014','COLLECTED as SRC-0109; TXN-0014. CO-0048 -$114,857.80 "Remove Landscaping packages from Omni Scope". Client-approved by David Wilson 2025-08-11. Counsel Status HOLD pending review.')
setrec(11,'R-0006',None,'COLLECTED as SRC-0109; TXN-0015 (CO-0028 -$3,061.27, approved 2025-04-03) and TXN-0016 (CO-0051 -$1,279.81, approved 2025-08-29), both client-approved by David Wilson. Combined -$4,341.08. Counsel Status HOLD pending review.')
print("R-0003..0006 updated")

wb.save(WB); print("SAVED")
