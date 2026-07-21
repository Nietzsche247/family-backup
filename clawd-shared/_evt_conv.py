import sys
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map
from openpyxl import load_workbook
M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'
s=load_json_strict(M+r'\Wilson.schema.json'); wb=load_workbook(M+r'\Wilson.xlsx')
ws=wb['MASTER CHRONOLOGY']; cm=col_map(s,'MASTER CHRONOLOGY')
inv={v:k for k,v in cm.items()}
def dumprow(r):
    print('--- row %d ---'%r)
    for col in range(1, ws.max_column+1):
        v=ws.cell(row=r,column=col).value
        if v is None: continue
        sv=str(v)
        print('  %2d %-26s = %s'%(col, inv.get(col,'?'), (sv[:70] if not sv.startswith('=') else '{FORMULA} '+sv[:55])))
# find rows for sample events
want={'EVT-0001':None,'EVT-0011':None,'EVT-0014':None,'EVT-0018':None,'EVT-0019':None}
last=None
for r in range(1, ws.max_row+1):
    eid=ws.cell(row=r,column=1).value
    if eid and str(eid).startswith('EVT-'):
        last=(r,str(eid))
        if str(eid) in want: want[str(eid)]=r
for e,r in want.items():
    if r: dumprow(r)
print('LAST EVENT:', last, '| next empty row =', (last[0]+1 if last else '?'))
print()
for k in ['date_type','statement_class','dispute_status','transcription_state','source_location_state','authenticity','admissibility','corroboration','counsel_status','record_state','hot']:
    if k in s.get('enums',{}): print('ENUM',k,'=',s['enums'][k])
