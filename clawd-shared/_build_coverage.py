import sys, os, shutil
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map, set_literal
from openpyxl import load_workbook

M = r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'
WB = os.path.join(M, 'Wilson.xlsx'); SC = os.path.join(M, 'Wilson.schema.json')
assert not os.path.exists(os.path.join(M, '~$Wilson.xlsx')), "OPEN in Excel"
bk = os.path.join(M, 'Wilson_prewrite_backup_2026-07-20_COVERAGE.xlsx')
shutil.copy2(WB, bk); print("BACKUP", bk)

schema = load_json_strict(SC); wb = load_workbook(WB)
cov = wb['COVERAGE']; cm = col_map(schema, 'COVERAGE')

rows = [
 {'Data Area':'Email - custodian universe (9 mailboxes)','Source Type key':'Email','Declared Items':7954,'Expected From':'2023-09-01','Expected To':'2026-07-17','Preserved / Hashed':7954,'Processed':3,'Quarantined':63,'Unresolved':7888,'Rows Produced':3,'Notes':'9 custodian mailboxes (aaron 1630, michael 1502, construction 1414, martin 1015, christine.stewart 961, scott.culver 498, veronica.leyva 443, build 404, allysa.redmon 87), 21.7 GB. 7,954 occurrences deduped to 5,282 distinct documents by Message-ID (2,672 duplicate occurrences). 63 occurrences quarantined for privilege review (C:/Omni-Privilege/Wilson). Preserved in full with SHA-256, no deletion. Only the Tom!! thread (SRC-0101/0102/0103) processed into the record; the rest preserved but not yet reviewed.'},
 {'Data Area':'Email - Wilson on envelope (Set B, targeted review)','Source Type key':'Email','Declared Items':109,'Expected From':'2023-09-01','Expected To':'2026-07-17','Preserved / Hashed':109,'Processed':3,'Unresolved':106,'Rows Produced':3,'Notes':'Header-scoped: David Wilson on To/From/Cc. Privilege-safe by construction (Wilson is on it). CAVEAT: header-scoping MISSES forwards where Wilson is only in quoted text (e.g. Fwd: Tom!!) - those need targeted thread pulls. Set E (Wilson AND Delaney on same envelope) = 17 messages. Set C (Delaney on envelope) = 2,746 = Delaney full mailbox, NOT the Wilson matter, excluded from review target. The 3 processed messages fall within this set.'},
 {'Data Area':'Change Orders','Source Type key':'Change Order','Declared Items':56,'Expected From':'2025-02-14','Expected To':'2026-07-20','Preserved / Hashed':56,'Processed':56,'Unresolved':0,'Rows Produced':4,'Notes':'All 56 change orders (CO-0001..CO-0062) captured in grid export SRC-0109 (subtotal $538,179.12), reflected in R-0003. 4 material scope deducts entered as transactions TXN-0013..0016 (CO-0036 -$63,917.18, CO-0048 -$114,857.80, CO-0028 -$3,061.27, CO-0051 -$1,279.81; combined -$183,116.06), all client-approved by David Wilson. Signed detail PDFs saved but font-subset (OCR pending); not required.'},
 {'Data Area':'Estimate (original)','Source Type key':'Estimate','Declared Items':1,'Expected From':'2025-02-12','Expected To':'2025-02-12','Preserved / Hashed':1,'Processed':1,'Unresolved':0,'Rows Produced':1,'Notes':'SRC-0108 Buildertrend estimate export, Total price $253,837.70, locked by Christine Stewart 2025-02-12. Feeds R-0002 (delta vs contract $81,075.88).'},
 {'Data Area':'Contract (signed)','Source Type key':'Contract','Declared Items':1,'Expected From':'2025-02-10','Expected To':'2025-02-10','Preserved / Hashed':1,'Processed':1,'Unresolved':0,'Rows Produced':4,'Notes':'SRC-0100 signed DocuSign construction contract, Total Price $334,913.58. 4 payment-schedule transactions TXN-0001..0004. Feeds R-0001 (ACCEPTED FOR USE).'},
 {'Data Area':'Pleadings (court filings)','Source Type key':'Pleading','Declared Items':1,'Expected From':'2026-06-12','Expected To':'2026-06-12','Preserved / Hashed':1,'Processed':1,'Unresolved':0,'Rows Produced':26,'Notes':'SRC-0107 complaint C20264565 filed 2026-06-12. 18 claim elements + 8 alleged-amount transactions (TXN-0005..0012) extracted. Answer and later filings not yet filed/collected; declared grows with the docket.'},
 {'Data Area':'Termination & cancellation docs','Source Type key':'Correspondence','Declared Items':3,'Preserved / Hashed':3,'Processed':3,'Unresolved':0,'Rows Produced':3,'Notes':'SRC-0104 Mutual Cancellation Release (2024-11-12), SRC-0105 Tom Delaney Termination, SRC-0106 Tom signed write-up. Two carry no embedded date in source bytes.'},
 {'Data Area':'Invoices & balance','Source Type key':'Invoice','Declared Items':1,'Notes':'NOT yet collected (Buildertrend browser export in progress). Reported: summary "Remaining balance" $36,377.04 vs invoice-grid "Balance due" $51,752.62 (invoice 0007, punch-list 5%, due 2026-07-25). Feeds R-0008. Two figures disagree; both to be captured.'},
 {'Data Area':'Payments (client)','Source Type key':'Payment','Declared Items':9,'Notes':'NOT yet collected (Buildertrend browser export in progress, /app/InvoicePayments). Nine Wilson payments, reported total $760,467.41, QuickBooks-recorded (QB and Buildertrend sync). Feeds R-0007.'},
 {'Data Area':'Daily logs','Source Type key':'Buildertrend','Declared Items':394,'Notes':'NOT collected. 394 daily logs reported on the job (Matthew Nelson authored 85). Counter-timeline: post-termination logs bear on the pleaded breach date. Declared count reported, not yet verified by export.'},
 {'Data Area':'Client comments (Buildertrend)','Source Type key':'Chat','Declared Items':110,'Expected From':'2025-02-14','Expected To':'2026-06-22','Notes':'NOT collected. 110 comments Feb 14 2025 to Jun 22 2026 (participants incl. David Wilson, Tom Delaney, Michael Baker, Allysa Redmon, Matthew Nelson, Veronica Leyva). Client comms flow through Buildertrend Comments, not email (Messages inbox empty). Reported by browser recon.'},
 {'Data Area':'Calendars','Source Type key':'Other','Notes':'NOT scoped, none collected. Meetings (e.g. termination meeting 2025-05-09, site visits) would live here. Universe not yet declared.'},
]

assert cov.cell(row=6, column=cm['Declared Items']).value in (None, ''), "COVERAGE row 6 not empty"
for i, rd in enumerate(rows):
    r = 6 + i
    for field, val in rd.items():
        c = cov.cell(row=r, column=cm[field])
        if isinstance(val, int):
            c.value = val
        else:
            set_literal(c, val)
print(f"wrote {len(rows)} COVERAGE rows (6-{5+len(rows)})")

wb.save(WB); print("SAVED")
