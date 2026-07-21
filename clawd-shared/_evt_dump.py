import sys
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map
from openpyxl import load_workbook
WB=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1\Wilson.xlsx'
s=load_json_strict(r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1\Wilson.schema.json')
wb=load_workbook(WB); ws=wb['MASTER CHRONOLOGY']; cm=col_map(s,'MASTER CHRONOLOGY')
def g(r,k):
    v=ws.cell(row=r,column=cm[k]).value
    return '' if v is None else str(v)
print('EVENTS:')
for r in range(1, ws.max_row+1):
    eid=g(r,'Event ID')
    if not eid.startswith('EVT-'): continue
    print('%s | %s | %s'%(eid, g(r,'Event Date')[:12], g(r,'Event Text (released core)')[:78]))
    print('      src=%s | by=%s | state=%s | hash=%s | receipt=%s | batch=%s'%(
        g(r,'Source ID(s)')[:20], g(r,'Entered By')[:14], g(r,'Record State')[:12],
        (g(r,'Record Hash')[:10] or 'NONE'), (g(r,'Receipt ID')[:14] or 'NONE'), (g(r,'Batch / Ordinal')[:10] or 'NONE')))
# staging / landing depth
for t in ['STAGING','IMPORT LANDING']:
    w=wb[t]; nonempty=0
    for row in w.iter_rows(min_row=6):
        if any(c.value is not None for c in row): nonempty+=1
    print(t,'data rows past header:',nonempty)
