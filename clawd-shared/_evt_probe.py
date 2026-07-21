import sys
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map
from openpyxl import load_workbook
M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'
s=load_json_strict(M+r'\Wilson.schema.json'); wb=load_workbook(M+r'\Wilson.xlsx'); ws=wb['MASTER CHRONOLOGY']; c=col_map(s,'MASTER CHRONOLOGY')
inv={v:k for k,v in c.items()}
# find EVT-0030 row and last EVT
rows={}
last=None; nextrow=None
for r in range(6, ws.max_row+2):
    v=ws.cell(row=r,column=1).value
    if v: rows[v]=r; last=(v,r)
    elif nextrow is None and r>6: nextrow=r; break
print('EVT-0030 at row', rows.get('EVT-0030'), '| last', last, '| next empty row', nextrow)
print('\n--- EVT-0030 full field pattern ---')
r=rows['EVT-0030']
for col in range(1,32):
    val=ws.cell(row=r,column=col).value
    print('  %-26s = %r'%(inv.get(col,'col%d'%col), val))
