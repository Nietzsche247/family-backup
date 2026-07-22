import os, sys, zipfile, hashlib
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map
from openpyxl import load_workbook
M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'; pres=M+r'\preserved'
def sha(p):
    h=hashlib.sha256()
    with open(p,'rb') as f:
        for c in iter(lambda:f.read(1024*1024),b''): h.update(c)
    return h.hexdigest()
# manifest
s=load_json_strict(M+r'\Wilson.schema.json'); wb=load_workbook(M+r'\Wilson.xlsx'); si=wb['SOURCE INDEX']; c=col_map(s,'SOURCE INDEX')
L=['WILSON v OMNI POOL BUILDERS - EVIDENCE SOURCE MANIFEST','Case: Pima County Superior Court C20264565','Generated: 2026-07-21  |  Working evidence set (attorney verification pending)','',
   'SOURCE ID | TYPE | SHA-256 (first 16) | WORKING COPY | DESCRIPTION','']
n=0
for r in range(6, si.max_row+1):
    sid=si.cell(row=r,column=c['Source ID']).value
    if not sid: continue
    n+=1
    L.append(' | '.join([str(sid), str(si.cell(row=r,column=c['Type']).value or '-'),
        (str(si.cell(row=r,column=c['SHA-256']).value or '')[:16] or '-'),
        str(si.cell(row=r,column=c['Working Copy (filename)']).value or '-'),
        str(si.cell(row=r,column=c['Description']).value or '')[:90]]))
L.append(''); L.append('Total sources: %d'%n)
open(M+r'\SOURCE_MANIFEST.txt','w',encoding='utf-8').write('\n'.join(L))
# handoff note
HN="""WILSON v OMNI POOL BUILDERS - COUNSEL HANDOFF NOTE
Case: Pima County Superior Court C20264565
Prepared: 2026-07-21  |  Working evidence set - attorney verification required

DEADLINE (COUNSEL MUST INDEPENDENTLY CONFIRM IMMEDIATELY):
Complaint filed 2026-06-12; service reportedly 2026-07-02; answer reportedly due
2026-07-22. Counsel must independently confirm immediately. This package does not
calculate or rely on any deadline.

FULL VAULT CHAT CONTAINER:
Full SRC-0130 Google Vault container preserved separately and available through
secure transfer. (Approx. 1.39 GB; not embedded in this bundle due to size.
Google MD5 and internal SHA-256 are recorded in the workbook SOURCE INDEX;
SRC-0131 is the extracted warning-meeting subset, included here.)

WHAT THIS IS:
A clerical evidence-organization work product for counsel to verify. Every fact
links to a preserved, hash-verified source. Counsel Status is UNREVIEWED
throughout by design; this package makes no legal-sufficiency determinations.

CONTENTS:
- Wilson.xlsx                        the evidence workbook (chronology, sources,
                                     claims/defenses, transactions, reconciliation)
- Wilson_Counsel_Strategy_v3.xlsx    current strategy summary
- preserved/                         all preserved source files
- SOURCE_MANIFEST.txt                source list with SHA-256 hashes

KNOWN OPEN ITEMS (flagged for counsel):
1. EVT-0039 (final-payment text on or about 2026-04-27) rests on Michael Baker's
   account in SRC-0123; the primary text screenshot is not yet preserved.
2. Two balance figures disagree: $51,752.62 (invoice grid) vs $36,377.04
   (job-remaining). The $15,375.58 delta is unreconciled; Christine Stewart to
   explain.
3. Workbook edits were direct-entry, not the hash-chained pipeline; change
   integrity rests on the external audits (omni_audit and omni_single_entry, both
   0 issues) plus dated backups. The sources themselves are all SHA-256 verified.
4. SRC-0112 (job-balance screenshot) not captured; the figure is documented in
   RECONCILIATION R-0008.
"""
open(M+r'\HANDOFF_NOTE.txt','w',encoding='utf-8').write(HN)
# build zip
bundle=M+r'\WILSON_evidence_bundle_2026-07-21.zip'
if os.path.exists(bundle): os.remove(bundle)
included=[]
with zipfile.ZipFile(bundle,'w',zipfile.ZIP_DEFLATED) as z:
    z.write(M+r'\Wilson.xlsx','Wilson.xlsx'); included.append('Wilson.xlsx')
    z.write(M+r'\HANDOFF_NOTE.txt','HANDOFF_NOTE.txt'); included.append('HANDOFF_NOTE.txt')
    z.write(M+r'\SOURCE_MANIFEST.txt','SOURCE_MANIFEST.txt'); included.append('SOURCE_MANIFEST.txt')
    v3=M+r'\Wilson_Counsel_Strategy_v3.xlsx'
    if os.path.exists(v3): z.write(v3,'Wilson_Counsel_Strategy_v3.xlsx'); included.append('Wilson_Counsel_Strategy_v3.xlsx')
    for root,dirs,files in os.walk(pres):
        for f in files:
            fp=os.path.join(root,f); z.write(fp,'preserved/'+os.path.relpath(fp,pres).replace('\\','/')); included.append('preserved/'+f)
# verify no stale strategy inside
with zipfile.ZipFile(bundle) as z: names=z.namelist()
stale=[x for x in names if 'strategy' in x.lower() and 'v3' not in x.lower()]
print('=== COUNSEL BUNDLE REBUILT ===')
print('bundle:', bundle)
print('total files in bundle:', len(names))
print('bundle size: %.1f MB'%(os.path.getsize(bundle)/1e6))
print('Wilson.xlsx SHA-256   :', sha(M+r'\Wilson.xlsx'))
print('strategy v3 SHA-256   :', sha(v3) if os.path.exists(v3) else 'MISSING')
print('stale strategy in zip :', stale if stale else 'NONE (confirmed removed)')
print('sources in manifest   :', n)
