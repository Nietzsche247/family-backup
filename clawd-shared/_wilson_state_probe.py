# read-only state probe of the live Wilson matter workbook
# does NOT modify the file. opens data_only=False so we can see formula vs static.
import openpyxl, os

P = r"C:\North_Star_Projects\WILSON_MATTER_WORKING_v1\Wilson.xlsx"
wb = openpyxl.load_workbook(P, data_only=False, read_only=True)

print("=== SHEETS (formula vs static) ===")
for ws in wb.worksheets:
    mr, mc = ws.max_row, ws.max_column
    formula = 0
    static_vals = 0
    data_rows = 0
    for row in ws.iter_rows():
        rowhas = False
        for c in row:
            v = c.value
            if v is None:
                continue
            if isinstance(v, str) and v.startswith("="):
                formula += 1
                rowhas = True
            else:
                static_vals += 1
                rowhas = True
        if rowhas:
            data_rows += 1
    print(f"{ws.title:<26} rows~{mr:<5} cols{mc:<3} | formulas={formula:<5} static={static_vals:<6} nonemptyRows={data_rows}")
wb.close()
