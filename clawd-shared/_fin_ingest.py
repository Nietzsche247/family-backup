import sys, os, shutil, hashlib
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map, set_literal
from openpyxl import load_workbook

M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'; WB=M+r'\Wilson.xlsx'; SC=M+r'\Wilson.schema.json'
FIN=M+r'\preserved\financial'; DL=r'C:\Users\aaron\Downloads'
assert not os.path.exists(M+r'\~$Wilson.xlsx'), "OPEN in Excel"
def cp(src,dn):
    d=os.path.join(FIN,dn); shutil.copy2(os.path.join(DL,src),d)
    return d, hashlib.sha256(open(d,'rb').read()).hexdigest()
d110,h110=cp('Invoices.xlsx','SRC-0110_owner-invoices-grid_WILSON.xlsx')
d111,h111=cp('InvoicePayments.xlsx','SRC-0111_invoice-payments-grid_WILSON.xlsx')
print('SRC-0110 sha',h110); print('SRC-0111 sha',h111)
for p in (d110,d111):
    w=load_workbook(p); ws=w.active; print(os.path.basename(p),'->',ws.max_row,'rows x',ws.max_column,'cols')
bk=M+r'\Wilson_prewrite_backup_2026-07-20_FIN.xlsx'; shutil.copy2(WB,bk); print('BACKUP',bk)
s=load_json_strict(SC); wb=load_workbook(WB); si=wb['SOURCE INDEX']; c=col_map(s,'SOURCE INDEX')
r=None
for rr in range(6,si.max_row+2):
    if not si.cell(row=rr,column=c['Source ID']).value: r=rr; break
assert r==20,'expected row 20 got %s'%r
def put(row,k,v): set_literal(si.cell(row=row,column=c[k]),v)
data={20:{'Source ID':'SRC-0110','SHA-256':h110,'Doc Date':'2026-07-20','Date Type':'CAPTURE',
 'Date Basis':'Export generated 2026-07-20 (Buildertrend Owner Invoices grid export).','Type':'Invoice',
 'Description':'Buildertrend Owner Invoices grid - Wilson job 38668322 (8 invoices 0001-0008; Total $812,220.03; Paid $760,467.41; Balance Due $51,752.62)',
 'Custodian':'Omni Pool Builders (Buildertrend job 38668322)','Native (relative URI)':'dir:staging_financial/Invoices_WILSON_grid_2026-07-20.xlsx',
 'Working Copy (filename)':'SRC-0110_owner-invoices-grid_WILSON.xlsx',
 'Acquisition Method':'A1 administrative export; Buildertrend Owner Invoices Export to Excel (POST /api/OwnerInvoices/ExportToExcel 200); Wilson job 38668322; exported 2026-07-20. Downloaded as Invoices.xlsx.',
 'Received Date':'2026-07-20','Disposition':'PRESERVED',
 'Notes':'Grid totals: Total price $812,220.03 | Amount paid $760,467.41 | Balance Due $51,752.62. 8 invoices (0001-0008). Invoice 0007 = "5% Payment Due at Completion of Punch List", Pending/Sent, $51,752.62 = the entire grid Balance Due; grid Due date shows 2026-07-20 (a prior note said 2026-07-25 - reconcile). Feeds R-0008. Job-level "Remaining balance" $36,377.04 is a separate screen-capture exhibit (balance bar, no export button).'},
 21:{'Source ID':'SRC-0111','SHA-256':h111,'Doc Date':'2026-07-20','Date Type':'CAPTURE',
 'Date Basis':'Export generated 2026-07-20 (Buildertrend Invoice Payments grid export).','Type':'Payment',
 'Description':'Buildertrend Invoice Payments grid - Wilson job (9 payments; total $760,467.41; all QuickBooks-synced)',
 'Custodian':'Omni Pool Builders (Buildertrend job 38668322)','Native (relative URI)':'dir:staging_financial/InvoicePayments_WILSON_grid_2026-07-20.xlsx',
 'Working Copy (filename)':'SRC-0111_invoice-payments-grid_WILSON.xlsx',
 'Acquisition Method':'A1 administrative export; Buildertrend Invoice Payments Export to Excel (POST /api/InvoicePayments/ExportToExcel 200); Wilson job 38668322; exported 2026-07-20. Downloaded as InvoicePayments.xlsx.',
 'Received Date':'2026-07-20','Disposition':'PRESERVED',
 'Notes':'9 payments, all Recorded by/method = QuickBooks, all Paid; total $760,467.41. 9 payments vs 8 invoices reconciles because invoices 0001 and 0006 each took two partial payments. QuickBooks is the operative ledger; this mirrors the QB-synced records. Feeds R-0007.'}}
for rr,d in data.items():
    for k,v in d.items(): put(rr,k,v)
    print('added',d['Source ID'],'row',rr)
rc=wb['RECONCILIATION']; rcm=col_map(s,'RECONCILIATION'); nk=rcm['Notes']
assert rc.cell(row=12,column=1).value=='R-0007' and rc.cell(row=13,column=1).value=='R-0008'
r7=('COLLECTED 2026-07-20 as SRC-0111 (Buildertrend Invoice Payments grid export; POST /api/InvoicePayments/ExportToExcel 200; 9 payments; total $760,467.41; all Recorded by/method = QuickBooks; all Paid). '
 'Breakdown: inv0001 2025-02-05 $6,256.07; inv0001 2025-03-28 $10,887.32; inv0006 2025-03-28 $773.49; inv0002 2025-04-15 $111,431.06; inv0003 2025-05-10 $214,292.43; '
 'inv0008 2025-05-05 $37,430.11; inv0004 2025-06-05 $165,217.84; inv0005 2025-10-30 $166,320.60; inv0006 2025-11-01 $47,858.49. '
 '9 payments vs 8 invoices reconciles (invoices 0001 and 0006 each took two partial payments). QuickBooks remains the operative ledger; export mirrors QB-synced records. Ready for counsel to move off HOLD.')
r8=('BOTH balances now sourced. Invoice-grid "Balance Due" $51,752.62 COLLECTED as SRC-0110 (Buildertrend Owner Invoices export; 8 invoices; Total $812,220.03; Paid $760,467.41). '
 'Job summary "Remaining balance" $36,377.04 (job running total $796,844.45 minus payments $760,467.41) remains a SCREEN-CAPTURE exhibit - the balance bar has no export button; full-page screenshot pending. '
 'The two disagree by $15,375.58 (invoice-grid Total $812,220.03 vs job running total $796,844.45); Christine to explain the gap. Invoice 0007 (5% punch-list retention, $51,752.62, Pending/Sent) = the entire invoice-grid balance due; '
 'grid Due date shows 2026-07-20 (a prior note said 2026-07-25 - reconcile). Do not present one balance without the other. NOT PLED in the complaint. Ready for counsel review.')
set_literal(rc.cell(row=12,column=nk),r7); set_literal(rc.cell(row=13,column=nk),r8)
print('R-0007 + R-0008 notes updated (Counsel Status left HOLD)')
cv=wb['COVERAGE']
cv.cell(row=13,column=7).value=1; cv.cell(row=13,column=12).value=0; cv.cell(row=13,column=13).value=8
cv.cell(row=14,column=7).value=9; cv.cell(row=14,column=12).value=0; cv.cell(row=14,column=13).value=9
set_literal(cv.cell(row=13,column=16),'SRC-0110 Buildertrend invoices grid (8 invoices). Balance Due $51,752.62 = invoice 0007. Job "remaining balance" $36,377.04 is a separate screen exhibit.')
set_literal(cv.cell(row=14,column=16),'SRC-0111 Buildertrend payments grid (9 payments, $760,467.41, QuickBooks-synced).')
print('COVERAGE invoices+payments -> processed')
wb.save(WB); print('SAVED')
