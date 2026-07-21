import sys
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from openpyxl import load_workbook
M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'
wb=load_workbook(M+r'\Wilson.xlsx')
print('%-22s %6s %5s %7s  %s'%('SHEET','maxR','maxC','dataRows','first data value (col A/B)'))
print('-'*95)
for ws in wb.worksheets:
    mr,mc=ws.max_row,ws.max_column
    # data region: rows 6+ (most sheets have header at row 5). count rows with any non-empty cell
    data=0; firstval=''
    for r in range(6, mr+1):
        rowvals=[ws.cell(row=r,column=col).value for col in range(1,min(mc,8)+1)]
        nonempty=[v for v in rowvals if v not in (None,'')]
        if nonempty:
            data+=1
            if not firstval:
                a=ws.cell(row=r,column=1).value; b=ws.cell(row=r,column=2).value
                firstval=str(a or b or nonempty[0])[:40]
    tag=''
    if data==0: tag='  <-- EMPTY'
    elif data<=3: tag='  <-- thin'
    print('%-22s %6d %5d %7d  %s%s'%(ws.title, mr, mc, data, firstval, tag))
