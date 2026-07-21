import os, sys, zipfile
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map
from openpyxl import load_workbook
M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'
pres=M+r'\preserved'
# size guard
tot=0; cnt=0
for root,dirs,files in os.walk(pres):
    for f in files:
        tot+=os.path.getsize(os.path.join(root,f)); cnt+=1
print('preserved/ = %d files, %.1f MB'%(cnt, tot/1e6))
if tot > 500e6:
    print('TOO LARGE - not zipping, check for stray large files'); sys.exit()
# manifest from SOURCE INDEX
s=load_json_strict(M+r'\Wilson.schema.json'); wb=load_workbook(M+r'\Wilson.xlsx'); si=wb['SOURCE INDEX']; c=col_map(s,'SOURCE INDEX')
L=['WILSON v OMNI POOL BUILDERS - EVIDENCE SOURCE MANIFEST','Case: Pima County Superior Court C20264565','Generated: 2026-07-21  |  Working evidence set (attorney verification pending)','',
   'Source files are under preserved/. The workbook Wilson.xlsx carries the full SOURCE INDEX, chronology, claims/defenses, transactions and reconciliation.','',
   'SOURCE ID | TYPE | SHA-256 (first 16) | WORKING COPY | DESCRIPTION','']
n=0
for r in range(6, si.max_row+1):
    sid=si.cell(row=r,column=c['Source ID']).value
    if not sid: continue
    n+=1
    L.append(' | '.join([str(sid), str(si.cell(row=r,column=c['Type']).value or '-'),
        (str(si.cell(row=r,column=c['SHA-256']).value or '')[:16] or '-'),
        str(si.cell(row=r,column=c['Working Copy (filename)']).value or '-'),
        str(si.cell(row=r,column=c['Description']).value or '')[:90]]))
L.append(''); L.append('Total sources: %d'%n)
open(M+r'\SOURCE_MANIFEST.txt','w',encoding='utf-8').write('\n'.join(L))
print('manifest: %d sources'%n)
# zip
bundle=M+r'\WILSON_evidence_bundle_2026-07-21.zip'
with zipfile.ZipFile(bundle,'w',zipfile.ZIP_DEFLATED) as z:
    z.write(M+r'\Wilson.xlsx','Wilson.xlsx')
    z.write(M+r'\SOURCE_MANIFEST.txt','SOURCE_MANIFEST.txt')
    sc=M+r'\Wilson_Counsel_Strategy.xlsx'
    if os.path.exists(sc): z.write(sc,'Wilson_Counsel_Strategy_FROZEN.xlsx')
    for root,dirs,files in os.walk(pres):
        for f in files:
            fp=os.path.join(root,f)
            z.write(fp, 'preserved/'+os.path.relpath(fp,pres).replace('\\','/'))
print('BUNDLE:', bundle, '%.1f MB'%(os.path.getsize(bundle)/1e6))
