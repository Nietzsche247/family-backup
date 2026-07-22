import sys, json
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from openpyxl import load_workbook
M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'
wb=load_workbook(M+r'\Wilson.xlsx')
ws=wb['RECONCILIATION']
# human-input columns (1-based): A1 B2 C3 D4 F6 H8 J10 N14 O15
HUMAN={'A':1,'B':2,'C':3,'D':4,'F':6,'H':8,'J':10,'N':14,'O':15}
recs=[]
for r in range(6,18):   # R-0001..R-0012 at rows 6-17
    rec={'_srcrow':r}
    for lab,col in HUMAN.items():
        v=ws.cell(row=r,column=col).value
        rec[lab]= '' if v is None else v
    recs.append(rec)
json.dump(recs, open(r'C:\Users\aaron\clawd-shared\_recon_data.json','w',encoding='utf-8'), indent=1, ensure_ascii=False)
print('captured %d records to _recon_data.json'%len(recs))
for rec in recs:
    print(' %s | %s | %s'%(rec['A'], rec['B'], str(rec['C'])[:40]))
    if rec['D'] or rec['F'] or rec['H'] or rec['J']:
        print('      D=%s F=%s H=%s J=%s N=%s'%(rec['D'],rec['F'],rec['H'],rec['J'],rec['N']))
# full E formula from row 9 for tx_last
print('\nrow9 E formula (full):')
print(' ', ws.cell(row=9,column=5).value)
print('row9 M formula (full):')
print(' ', ws.cell(row=9,column=13).value)
# check any human values sitting in rows 18-20 (should be none)
print('\nrows 18-20 human cols (should be empty):')
for r in range(18,21):
    vals={lab:ws.cell(row=r,column=col).value for lab,col in HUMAN.items() if ws.cell(row=r,column=col).value not in (None,'')}
    print('  r%d %s'%(r, vals or '(empty)'))
