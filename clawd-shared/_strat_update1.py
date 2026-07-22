import os, shutil
from openpyxl import load_workbook
M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'
SRC=M+r'\Wilson_Counsel_Strategy_v3.xlsx'
OUT=M+r'\Wilson_Counsel_Strategy_FINAL_2026-07-21.xlsx'
WSHA='c9db4204bc3687d67cb582f2251a3f4a8662b8f11025832842c122a2e151c302'
shutil.copy2(SRC,OUT)
wb=load_workbook(OUT)
# report any tables (so appending a row cannot break one)
for sh in wb.sheetnames:
    for t in wb[sh].tables.values(): print('TABLE %s on %s ref=%s'%(t.name,sh,t.ref))

# ---------- EXECUTIVE ----------
ex=wb['EXECUTIVE']
ex['A4']=('Current source workbook: Wilson.xlsx \u2022 SHA-256 '+WSHA+' \u2022 snapshot 2026-07-21 post-repair '
          '\u2022 supersedes the pre-repair workbook 85eb136f77e8af8f683989bb436125abe90a4f8d01cf13afcd615449531500d0 from which strategy v3 was generated')
ex['B7']=('Useful and substantial as a clerical orientation package. The RECONCILIATION structural repair is complete and every technical check now passes '
          '(omni_audit 0 issues, omni_single_entry 0 issues, native Microsoft Excel recalculation 0 formula errors). The content itself remains DRAFT and '
          'UNREVIEWED and the collection is targeted rather than exhaustive. Sendable to counsel with the open proof listed in this file.')
ex['D19']='CD-0015, CD-0018 and CD-0019 links refreshed 2026-07-21; CD-0004, CD-0009 and CD-0017 still have no supporting Event IDs'
ex['D22']='Structural repair completed 2026-07-21: header at row 8, R-0001 to R-0012 in table data rows 9-20, generated formulas restored'
print('EXECUTIVE updated')

# ---------- TECHNICAL READINESS ----------
tr=wb['TECHNICAL READINESS']
tr['A5']=('Technical structure now passes on the final workbook: omni_audit 0 issues, omni_single_entry 0 issues, RECONCILIATION structure repaired, and a '
          'native Microsoft Excel open/recalculate/save returning 0 formula errors. This remains a clerical package, not a verified factual record. Confirm '
          'the response deadline independently of this file and disclose the unresolved review universe. The strategy workbook accompanies it as an expressly '
          'unverified attorney-review draft.')
tr['C8']='PASS'
tr['D8']='omni_audit 14.4.0 returned 0 issues on the final workbook (SHA c9db4204).'
tr['E8']='Structural audit only. It does not verify factual accuracy, transcription, or legal sufficiency.'
tr['C9']='PASS'
tr['D9']=('omni_single_entry returned 0 issues on the final workbook. The prior 15 SE-WB-03 findings are resolved by the RECONCILIATION header repair. '
          'Role counts: RELATIONSHIP 34, SYSTEM 14, HUMAN_DECISION 19, CANONICAL_INPUT 180, GENERATED 37.')
tr['E9']='No action. Rerun after any further structural edit.'
tr['C10']='PASS'
tr['D10']=('Native Microsoft Excel open/recalculate/save (CalculateFullRebuild) returned 0 formula error cells across all sheets, and correct cached values '
           'were written into the file.')
tr['E10']='No action. Missing inputs still yield intentional blanks by design.'
tr['C12']='PARTIAL'
tr['D12']=('19 entries. CD-0015, CD-0018 and CD-0019 event links were refreshed on 2026-07-21 from the final workbook. CD-0004, CD-0009 and CD-0017 still '
           'have no supporting Event IDs.')
tr['E12']='Collect support for the remaining predicates. Status and Counsel Status were not changed; counsel decides legal status.'
tr['B18']='Existing strategy v2 and v3'
tr['C18']='SUPERSEDED'
tr['D18']=('v2 reflects 8 sources, 19 events, 12 transactions and 18 claims. v3 was generated from the pre-repair workbook (SHA 85eb136f) and its TECHNICAL '
           'READINESS and CLAIM MAP no longer match the final workbook.')
tr['E18']='Ship only Wilson_Counsel_Strategy_FINAL_2026-07-21.xlsx. v2 and v3 are excluded from the counsel bundle.'
# append the RECONCILIATION structure result as entry 13
tr['A20']=13
tr['B20']='RECONCILIATION structure'
tr['C20']='PASS'
tr['D20']=('Header restored to row 8 with all 15 schema columns; R-0001 through R-0012 placed in table data rows 9-20 in order; generated formulas restored '
           'in columns E, G, I, K, L and M; tblRecon retained at A8:O208; human inputs and notes preserved verbatim.')
tr['E20']='No action. No approval or payment IDs were invented; those inputs remain blank for counsel.'
for c in 'ABCDE':
    tr['%s20'%c]._style = tr['%s19'%c]._style
print('TECHNICAL READINESS updated (+ entry 13 RECONCILIATION structure)')
wb.save(OUT); print('saved',os.path.basename(OUT))
