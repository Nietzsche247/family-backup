from openpyxl import load_workbook
wb = load_workbook(r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1\Wilson.xlsx', read_only=True)
cov = wb['COVERAGE']
def num(v):
    return v if isinstance(v,(int,float)) else 0
print(f"{'DATA AREA':<46} {'DECL':>6} {'PROC':>5} {'UNRES':>6} {'CHK':>5}  STATUS")
print("-"*90)
for r in range(6,18):
    A=cov.cell(r,1).value
    if not A: continue
    C=cov.cell(r,3).value
    G,H,I,J,K,L=[num(cov.cell(r,col).value) for col in (7,8,9,10,11,12)]
    ctype = type(C).__name__
    if C in (None,''):
        status="DECLARE UNIVERSE"; chk=""
    else:
        s=G+H+I+J+K+L
        chk=num(C)-s
        if s==0: status="NOT STARTED"
        elif chk==0 and L==0: status="RECONCILED"
        else: status="UNRECONCILED"
    decl = f"{C}({ctype[:3]})" if C not in (None,'') else "-"
    print(f"{A[:46]:<46} {str(decl):>6} {G:>5} {L:>6} {str(chk):>5}  {status}")
wb.close()
