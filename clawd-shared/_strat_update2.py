import sys
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map
from openpyxl import load_workbook
M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'
OUT=M+r'\Wilson_Counsel_Strategy_FINAL_2026-07-21.xlsx'
# read authoritative CLAIMS & DEFENSES
s=load_json_strict(M+r'\Wilson.schema.json'); src=load_workbook(M+r'\Wilson.xlsx', data_only=False)
cd=src['CLAIMS & DEFENSES']; c=col_map(s,'CLAIMS & DEFENSES')
recs=[]
for r in range(6, cd.max_row+1):
    cid=cd.cell(row=r,column=c['ID']).value
    if not cid: continue
    recs.append({'id':cid,
      'side':cd.cell(row=r,column=c['Side']).value or '',
      'claim':cd.cell(row=r,column=c['Claim / Defense']).value or '',
      'elem':cd.cell(row=r,column=c['Element or Predicate']).value or '',
      'sup':cd.cell(row=r,column=c['Supporting Event IDs']).value or '',
      'con':cd.cell(row=r,column=c['Contrary Event IDs']).value or '',
      'st':cd.cell(row=r,column=c['Status']).value or '',
      'cs':cd.cell(row=r,column=c['Counsel Status']).value or ''})
print('read %d CD records from final workbook'%len(recs))
wb=load_workbook(OUT); cm=wb['CLAIM MAP']
cm['A2']='Regenerated from the final workbook 2026-07-21; counsel determines legal sufficiency'
cm['A4']='%d CURRENT ENTRIES'%len(recs)
for i,rec in enumerate(recs):
    r=6+i
    cm.cell(row=r,column=1).value=rec['id']; cm.cell(row=r,column=2).value=rec['side']
    cm.cell(row=r,column=3).value=rec['claim']; cm.cell(row=r,column=4).value=rec['elem']
    cm.cell(row=r,column=5).value=rec['sup'] or None
    cm.cell(row=r,column=6).value=rec['con'] or None
    cm.cell(row=r,column=7).value=rec['st']; cm.cell(row=r,column=8).value=rec['cs']
print('CLAIM MAP rows 6-%d regenerated'%(5+len(recs)))
rowof={rec['id']:6+i for i,rec in enumerate(recs)}
cm.cell(row=rowof['CD-0015'],column=9).value=('RESOLVED 2026-07-21: EVT-0041 (2025-05-19 warning meeting; SRC-0131 contemporaneous chat plus SRC-0118 Wilson admission) '
  'added alongside EVT-0014 and EVT-0016. Counsel decides which notice event to lead with.')
cm.cell(row=rowof['CD-0018'],column=9).value=('RESOLVED 2026-07-21: EVT-0022 through EVT-0025 linked as client-approved scope deductions; TXN-0013 through TXN-0016 noted.')
cm.cell(row=rowof['CD-0019'],column=9).value=('PARTIALLY RESOLVED 2026-07-21: EVT-0041 linked as support for the express-warning assertion. Still unlinked and flagged for counsel: '
  'repeated banning of Delaney from the job; the written cease-and-desist (preserved as SRC-0115, no chronology event); Wilson\u2019s independent engagement of Delaney '
  '(partial support SRC-0103 plus the change-order scope-removal events).')
print('CD-0015 / CD-0018 / CD-0019 refresh notes updated')
for cid in ['CD-0015','CD-0018','CD-0019']:
    r=rowof[cid]
    print('  %s row%d  SUP=%s | STATUS=%s | COUNSEL=%s'%(cid,r,cm.cell(row=r,column=5).value,cm.cell(row=r,column=7).value,cm.cell(row=r,column=8).value))
# ---------- OPEN PROOF ----------
op=wb['OPEN PROOF']
op['B19']=('Wilson.xlsx with the RECONCILIATION structural repair complete, SOURCE_MANIFEST.txt, HANDOFF_NOTE.txt, the preserved source files, '
           'and this strategy workbook (Wilson_Counsel_Strategy_FINAL_2026-07-21.xlsx).')
op['B22']=('Wilson_Counsel_Strategy.xlsx (frozen 2026-07-17), v2, and v3. v3 was generated from the pre-repair workbook, so its technical sections contradict '
           'the final workbook. All are superseded by Wilson_Counsel_Strategy_FINAL_2026-07-21.xlsx and are excluded from the counsel bundle.')
op['B23']=('Reported 2026-07-22 response date; balance mismatch; missing primary 4/27 text; 106 direct Wilson emails and 110 Buildertrend comments not yet processed.')
print('OPEN PROOF updated (r19, r22, r23)')
wb.save(OUT); print('SAVED', OUT)
