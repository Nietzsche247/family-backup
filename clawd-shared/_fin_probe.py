import sys, os
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map
from openpyxl import load_workbook
M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'
DL=r'C:\Users\aaron\Downloads'
# 1) downloads
for f in ['OwnerInvoices.xlsx','InvoicePayments.xlsx']:
    p=os.path.join(DL,f); print('DL', f, 'EXISTS' if os.path.exists(p) else 'MISSING', os.path.getsize(p) if os.path.exists(p) else '')
# 2) preserved dirs
pr=os.path.join(M,'preserved'); print('preserved/ subdirs:', [d for d in os.listdir(pr) if os.path.isdir(os.path.join(pr,d))])
s=load_json_strict(M+r'\Wilson.schema.json'); wb=load_workbook(M+r'\Wilson.xlsx')
si=wb['SOURCE INDEX']; c=col_map(s,'SOURCE INDEX')
def g(ws,r,k,cm): v=ws.cell(row=r,column=cm[k]).value; return v
print('--- SRC-0108/0109 convention ---')
for r in range(6, si.max_row+1):
    sid=str(si.cell(row=r,column=c['Source ID']).value or '')
    if sid in ('SRC-0108','SRC-0109'):
        for k in ['Type','Doc Date','Date Type','Date Basis','Native (relative URI)','Working Copy (filename)','Acquisition Method','Disposition']:
            print('  %s %-24s: %r'%(sid,k,g(si,r,k,c)))
# 3) RECONCILIATION
print('--- RECONCILIATION ---')
rc=wb['RECONCILIATION']; rcm=col_map(s,'RECONCILIATION'); print('COLS:',rcm)
idk=list(rcm.keys())[0]
for r in range(1, rc.max_row+1):
    v0=rc.cell(row=r,column=1).value
    if v0 and str(v0).startswith('R-'):
        print('  r%d:'%r, [ (k, rc.cell(row=r,column=cc).value) for k,cc in rcm.items() ])
# 4) COVERAGE financial rows
print('--- COVERAGE financial rows ---')
cv=wb['COVERAGE']; cvm=col_map(s,'COVERAGE')
for r in range(6, cv.max_row+1):
    da=str(cv.cell(row=r,column=1).value or '')
    if any(x in da.lower() for x in ['invoice','payment','balance','financ']):
        print('  r%d DataArea=%r Declared=%r Processed=%r Rows=%r Status=%r'%(r,da,cv.cell(row=r,column=3).value,cv.cell(row=r,column=7).value,cv.cell(row=r,column=13).value,cv.cell(row=r,column=15).value))
