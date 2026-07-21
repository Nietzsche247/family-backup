import sys, os
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map, set_literal
from openpyxl import load_workbook
M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'; WB=M+r'\Wilson.xlsx'; SC=M+r'\Wilson.schema.json'
assert not os.path.exists(M+r'\~$Wilson.xlsx'), "OPEN in Excel"
s=load_json_strict(SC)
# find Issue Tags ref -> is it multi?
def find_ref():
    sh=s.get('sheets',s)
    for key in ['MASTER CHRONOLOGY']:
        node=sh.get(key) if isinstance(sh,dict) else None
        if not node: continue
        cols=node.get('columns') or node.get('cols') or []
        for cdef in cols:
            nm=cdef.get('name') or cdef.get('header')
            if nm and 'Issue Tags' in nm: return cdef
    return None
ref=find_ref(); print('Issue Tags ref:', ref)
multi = bool(ref.get('multi')) if ref else True
delim=';' if multi else None
print('multi=',multi)
wb=load_workbook(WB); ch=wb['MASTER CHRONOLOGY']; cc=col_map(s,'MASTER CHRONOLOGY')
if multi:
    set_literal(ch.cell(row=44,column=cc['Issue Tags']),'PAY;DEM')
    set_literal(ch.cell(row=45,column=cc['Issue Tags']),'WARR;PAY')
else:
    set_literal(ch.cell(row=44,column=cc['Issue Tags']),'PAY')
    set_literal(ch.cell(row=45,column=cc['Issue Tags']),'WARR')
newE45=("Omni stated its warranty position to Wilson, conditioned on final payment. Per Michael Baker (SRC-0123): "
        "'If you make payment this week I will activate your warranty right away and work on resolving all issues.' "
        "Omni was already proceeding with some warranty items and triaged issues into still-scheduled, waiting-for-final-payment, "
        "and never-in-scope. Consistent with the 2026-07-17 letter (SRC-0122) declining full warranty, offering limited warranty.")
print('E45 len =',len(newE45))
set_literal(ch.cell(row=45,column=cc['Event Text (released core)']),newE45)
wb.save(WB); print('SAVED tags=%s'%('PAY;DEM / WARR;PAY' if multi else 'PAY / WARR'))
