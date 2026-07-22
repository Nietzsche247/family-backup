import sys, os, json, time
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from openpyxl import load_workbook
M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'
wb=load_workbook(M+r'\Wilson.xlsx')
print('=== SHEETS (%d) : data rows counted from row 6, col A ==='%len(wb.sheetnames))
for sh in wb.sheetnames:
    ws=wb[sh]; n=0; first=None
    for r in range(6, ws.max_row+1):
        v=ws.cell(row=r,column=1).value
        if v not in (None,''):
            n+=1
            if first is None: first=str(v)[:14]
    print('  %-22s rows=%-4d  first=%s'%(sh, n, first or '-'))
print('\n=== MATTER FOLDER ===')
for f in sorted(os.listdir(M)):
    p=os.path.join(M,f)
    if os.path.isdir(p):
        cnt=sum(len(files) for _,_,files in os.walk(p))
        sz=sum(os.path.getsize(os.path.join(r,x)) for r,_,fs in os.walk(p) for x in fs)
        print('  [DIR ] %-52s %4d files  %8.1f MB'%(f,cnt,sz/1e6))
    else:
        print('  [FILE] %-52s %8.2f MB  %s'%(f, os.path.getsize(p)/1e6, time.strftime('%Y-%m-%d %H:%M',time.localtime(os.path.getmtime(p)))))
print('\n=== preserved/ subfolders ===')
pres=M+r'\preserved'
for root,dirs,files in os.walk(pres):
    rel=os.path.relpath(root,pres)
    if files: print('  %-24s %d files'%(rel, len(files)))
