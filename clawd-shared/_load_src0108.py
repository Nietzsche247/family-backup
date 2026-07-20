import sys, os, shutil, datetime, hashlib
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map, set_literal
from openpyxl import load_workbook

MATTER = r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'
WB = os.path.join(MATTER, 'Wilson.xlsx')
SC = os.path.join(MATTER, 'Wilson.schema.json')

assert not os.path.exists(os.path.join(MATTER, '~$Wilson.xlsx')), "Wilson.xlsx OPEN in Excel - abort"

# 0) backup before any write
bk = os.path.join(MATTER, 'Wilson_prewrite_backup_2026-07-20_SRC0108.xlsx')
shutil.copy2(WB, bk)
print("BACKUP:", bk)

# 1) preserve working copy into preserved/financial
staging_native = os.path.join(MATTER, 'staging4_financial', 'WILSON_BT_EstimateReport_original-estimate_exported_2026-07-20.xls')
assert os.path.exists(staging_native), "staged native missing"
preserved_dir = os.path.join(MATTER, 'preserved', 'financial')
os.makedirs(preserved_dir, exist_ok=True)
working = os.path.join(preserved_dir, 'SRC-0108_original-estimate_WILSON-BT.xls')
shutil.copy2(staging_native, working)
sha = hashlib.sha256(open(working, 'rb').read()).hexdigest()
mt = datetime.datetime.fromtimestamp(os.path.getmtime(staging_native), tz=datetime.timezone.utc).isoformat()
print("WORKING COPY:", working)
print("SHA256:", sha)

schema = load_json_strict(SC)
wb = load_workbook(WB)

# 2) SOURCE INDEX -> SRC-0108 at row 14
si = wb['SOURCE INDEX']; cm = col_map(schema, 'SOURCE INDEX')
ROW = 14
assert si.cell(row=ROW, column=cm['Source ID']).value in (None, ''), "SOURCE INDEX row 14 not empty"
vals = {
 'Source ID': 'SRC-0108',
 'SHA-256': sha,
 'Date Basis': 'NO EMBEDDED DATE - Buildertrend Excel export contains no internal date. Estimate reportedly locked 2025-02-12 by Christine Stewart per prior record; not verifiable from these bytes.',
 'Type': 'Estimate',
 'Description': 'Buildertrend Estimate module export (Excel), Wilson job 38668322. 38 line items by cost code; Builder Cost = Client Price = $253,837.70 (original estimate).',
 'Custodian': 'aaron@omnipoolbuilders.com',
 'From / Author': 'Buildertrend (Omni builder acct 6769), system export',
 'Native (relative URI)': 'dir:staging4_financial/WILSON_BT_EstimateReport_original-estimate_exported_2026-07-20.xls',
 'Working Copy (filename)': 'SRC-0108_original-estimate_WILSON-BT.xls',
 'Acquisition Method': 'A1 administrative export; Buildertrend Estimate module Export to Excel; Wilson job 38668322; exported 2026-07-20.',
 'Received Date': '2026-07-20',
 'Disposition': 'PRESERVED',
 'Notes': 'filesystem mtime ' + mt + '. Line-item total $253,837.70 recomputed in code (Builder Cost and Client Price columns both sum to it). PROVENANCE CAVEAT: export bytes contain no embedded jobId; Wilson linkage rests on (a) total matching the reported Original Estimate to the cent and (b) custody - exported to Aaron Downloads 2026-07-20 11:26. Confirm export originated from jobId 38668322 before treating as settled.',
}
for f, v in vals.items():
    if f in cm:
        set_literal(si.cell(row=ROW, column=cm[f]), v)
print("SOURCE INDEX row 14 written: SRC-0108")

# 3) RECONCILIATION R-0002 note only (row 7); leave Counsel Status HOLD
rec = wb['RECONCILIATION']; rm = col_map(schema, 'RECONCILIATION')
assert rec.cell(row=7, column=rm['Line ID']).value == 'R-0002', "R-0002 not at row 7"
note = ('COLLECTED 2026-07-20 as SRC-0108 (Buildertrend estimate export). Original Estimate $253,837.70, recomputed '
        'from the export line items. Contract Total Price $334,913.58 (SRC-0100, TXN-0001). Delta = $81,075.88 '
        '(contract over estimate). Counsel Status left HOLD pending review; ready to move to ACCEPTED FOR USE. '
        'Caveat: export bytes carry no embedded jobId - confirm it came from jobId 38668322.')
set_literal(rec.cell(row=7, column=rm['Notes']), note)
print("R-0002 note updated")

wb.save(WB)
print("SAVED Wilson.xlsx")
