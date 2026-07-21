import sys, os, shutil, hashlib
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map, set_literal
from openpyxl import load_workbook
M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'; WB=M+r'\Wilson.xlsx'; SC=M+r'\Wilson.schema.json'; KEY=M+r'\preserved\keydocs'
assert not os.path.exists(M+r'\~$Wilson.xlsx'), "OPEN in Excel"
def H(fn): return hashlib.sha256(open(os.path.join(KEY,fn),'rb').read()).hexdigest()
bk=M+r'\Wilson_prewrite_backup_2026-07-21_PROMOTE.xlsx'; shutil.copy2(WB,bk); print('BACKUP',bk)
s=load_json_strict(SC); wb=load_workbook(WB)
# ---------- image sources SRC-0124..0129 (skip img_6 divider) ----------
si=wb['SOURCE INDEX']; c=col_map(s,'SOURCE INDEX')
r0=None
for rr in range(6,si.max_row+2):
    if not si.cell(row=rr,column=c['Source ID']).value: r0=rr; break
assert r0==27,'expected 27 got %s'%r0
IMGS=[('SRC-0124','SRC-0123_img_1.png',"Buildertrend final-payment invoice released to Wilson 2026-03-16 - the invoice Michael states Wilson never opened. Embedded in SRC-0123 point 1."),
 ('SRC-0125','SRC-0123_img_2.png',"Screenshot in SRC-0123 under 'Issues still scheduled for rectification' - warranty items Omni was proceeding to fix."),
 ('SRC-0126','SRC-0123_img_3.png',"Screenshot in SRC-0123 under 'Issues waiting for final payment' - warranty items gated on Wilson's final payment."),
 ('SRC-0127','SRC-0123_img_4.png',"Screenshot in SRC-0123 = the final invoice ('Here is the final invoice')."),
 ('SRC-0128','SRC-0123_img_5.png',"Screenshot in SRC-0123 under 'Never in our scope' - items Omni states were never in its scope."),
 ('SRC-0129','SRC-0123_img_7.png',"Screenshot in SRC-0123 under 'Issues still scheduled for rectification' - continuation of the scheduled-items list.")]
for i,(sid,fn,desc) in enumerate(IMGS):
    r=r0+i; d={'Source ID':sid,'SHA-256':H(fn),'Doc Date':'2026-07-20','Date Type':'CAPTURE',
     'Date Basis':'Screenshot embedded in the 2026-07-20 email (SRC-0123); capture date = send date; depicted content per the email label.',
     'Type':'Photo','Description':desc,'Custodian':'Aaron Baker (aaron mailbox) / Michael Baker',
     'From / Author':'Omni (Michael Baker)','To / Recipients':'David Wilson',
     'Native (relative URI)':'embedded image in gmail "Re: Warranties and job completion items" 2026-07-20 (SRC-0123)',
     'Working Copy (filename)':fn,'Acquisition Method':'Extracted from the native 2026-07-20 email (SRC-0123), downloaded from Gmail 2026-07-21.',
     'Received Date':'2026-07-21','Disposition':'PRESERVED','Notes':'Standalone exhibit extracted from SRC-0123 embedded screenshots (7 total; divider image not sourced).'}
    for k,v in d.items(): set_literal(si.cell(row=r,column=c[k]),v)
    print('src',sid,'row',r)
# ---------- promote events EVT-0038..0040 ----------
ch=wb['MASTER CHRONOLOGY']; cc=col_map(s,'MASTER CHRONOLOGY')
er=None
for rr in range(6,ch.max_row+2):
    if not ch.cell(row=rr,column=1).value: er=rr; break
assert er==43,'expected 43 got %s'%er
COMMON={'Witness (disclosure)':None,'Transcription State':'NOT CHECKED','Source Location State':'PRESERVED',
 'Authenticity / Foundation':'NOT ASSESSED','Admissibility':'NOT ASSESSED','Corroboration':'SINGLE SOURCE',
 'Counsel Status':'UNREVIEWED','Entered By':'aaron baker','Entered On':'2026-07-21','Record State':'DRAFT','Dispute Status':'UNDISPUTED FOR NOW'}
EV=[
 {'Event ID':'EVT-0038','Event Date':'2026-03-16','Date Type':'SENT','Qualifier':'on or about',
  'Event Text (released core)':"Omni released the final-payment invoice (invoice 0007, 5% completion retention) to Wilson via Buildertrend. Per Michael Baker (SRC-0123): 'We released the final payment to you via BT March 16, 2026... It looks like you never opened this invoice.' Screenshot of the released invoice preserved as SRC-0124. Rebuts Wilson's 2026-07-19 claim he received no remaining invoices.",
  'Statement Class':'SYSTEM RECORD','Attribution / Speaker':'Omni / Buildertrend (per Michael Baker)','Actors':'ACT-0005; ACT-0001',
  'Source ID(s)':'SRC-0123; SRC-0124','Locator':'SRC-0123 point 1; BT invoice screenshot = SRC-0124 (SRC-0123_img_1.png)','Issue Tags':'PAY','Hot':'Y',
  'Batch / Ordinal':'STAGING-PROMOTE-2026-07-21/1'},
 {'Event ID':'EVT-0039','Event Date':'2026-04-27','Date Type':'SENT','Qualifier':'on or about',
  'Event Text (released core)':"Omni requested final payment from Wilson by text message. Per Michael Baker (SRC-0123): 'We also requested Final payment Via text on April 27th 2026.' Primary text screenshot not yet preserved (to be pulled from Michael's phone / 714-269-6161); this event rests on Michael's contemporaneous account.",
  'Statement Class':'PARTY ASSERTION','Attribution / Speaker':'Michael Baker (email)','Actors':'ACT-0005; ACT-0001',
  'Source ID(s)':'SRC-0123','Locator':'SRC-0123 point 2 (Michael 2026-07-20). Primary text record pending (Michael phone / 714-269-6161).','Issue Tags':'PAY,DEM','Hot':None,
  'Batch / Ordinal':'STAGING-PROMOTE-2026-07-21/2'},
 {'Event ID':'EVT-0040','Event Date':'2026-07-20','Date Type':'SENT','Qualifier':None,
  'Event Text (released core)':"Omni stated its warranty position to Wilson, conditioned on final payment. Per Michael Baker (SRC-0123): 'If you make payment this week I will activate your warranty right away and work on resolving all issues.' Michael noted Omni was already proceeding with some warranty items (contractors scheduled) and triaged issues into still-scheduled-for-rectification (SRC-0125, SRC-0129), waiting-for-final-payment (SRC-0126), and never-in-scope (SRC-0128), plus the final invoice (SRC-0127). Consistent with Omni's 2026-07-17 position letter (SRC-0122) declining full contractual/warranty completion and offering limited warranty items.",
  'Statement Class':'DOCUMENT CONTENT','Attribution / Speaker':'Michael Baker (email)','Actors':'ACT-0005; ACT-0001',
  'Source ID(s)':'SRC-0123; SRC-0122; SRC-0121','Locator':'SRC-0123 (Michael 2026-07-20); SRC-0122 = 2026-07-17 position letter; SRC-0125/0126/0127/0128/0129 = embedded triage screenshots.','Issue Tags':'WARR,PAY','Hot':'Y',
  'Batch / Ordinal':'STAGING-PROMOTE-2026-07-21/3'},
]
for i,d in enumerate(EV):
    r=er+i; full=dict(COMMON); full.update(d)
    for k,v in full.items():
        if v is not None: set_literal(ch.cell(row=r,column=cc[k]),v)
    ch.cell(row=r,column=cc['Check (auto)']).value='=IF($A%d="","",TRIM(IF($B%d="","NO DATE ","")&IF($J%d="","NO SOURCE ","")&IF($L%d="","NO TAGS ","")))'%(r,r,r,r)
    ch.cell(row=r,column=cc['HotRank (auto)']).value='=IF($M%d="Y",COUNTIF($M$6:$M%d,"Y"),"")'%(r,r)
    print('event',d['Event ID'],'row',r)
# clear stale native-pending note on EVT-0030
r30=None
for rr in range(6,ch.max_row+1):
    if ch.cell(row=rr,column=1).value=='EVT-0030': r30=rr; break
set_literal(ch.cell(row=r30,column=cc['Locator']),'gmail thread 19f72526564ff643; forward msg 19f7b36f8054112b. Native .eml preserved as SRC-0118 (2026-07-21).')
print('EVT-0030 locator updated (row %d)'%r30)
# ---------- STAGING: mark 1,2,4 moved ----------
st=wb['STAGING']; stc=col_map(s,'STAGING')
def SM(row,evt): set_literal(st.cell(row=row,column=stc['Status']),'MOVED TO CHRONOLOGY'); \
    set_literal(st.cell(row=row,column=stc['Where to Look']),'Promoted to '+evt+' (2026-07-21).')
SM(6,'EVT-0038'); SM(7,'EVT-0039'); SM(9,'EVT-0040')
print('STAGING items 1,2,4 -> MOVED TO CHRONOLOGY (item 3 stays FOUND)')
wb.save(WB); print('SAVED')
