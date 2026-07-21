import sys, os, shutil
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map, set_literal
from openpyxl import load_workbook
M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'; WB=M+r'\Wilson.xlsx'; SC=M+r'\Wilson.schema.json'
assert not os.path.exists(M+r'\~$Wilson.xlsx'), "OPEN in Excel"
bk=M+r'\Wilson_prewrite_backup_2026-07-20_EVENTS.xlsx'; shutil.copy2(WB,bk); print('BACKUP',bk)
s=load_json_strict(SC); wb=load_workbook(WB); ws=wb['MASTER CHRONOLOGY']; c=col_map(s,'MASTER CHRONOLOGY')
CONST={'Transcription State':'NOT CHECKED','Source Location State':'PRESERVED','Authenticity / Foundation':'NOT ASSESSED',
 'Admissibility':'NOT ASSESSED','Counsel Status':'UNREVIEWED','Entered By':'aaron baker','Entered On':'2026-07-20','Record State':'DRAFT'}
EV=[
 dict(id='EVT-0020',date='2025-02-12',dt='EVENT',q='on or about',sc='SYSTEM RECORD',attr='Buildertrend Estimate export (SRC-0108)',act='ACT-0008',src='SRC-0108',tag='EST',hot='',corr='SINGLE SOURCE',
   text='Original Buildertrend estimate locked at $253,837.70 by Christine Stewart. Contract Total Price later $334,913.58, a delta of $81,075.88 over the estimate.',
   loc='Recomputed from export line items. Date 2025-02-12 per prior record; export bytes carry no embedded date.'),
 dict(id='EVT-0021',date='2025-05-09',dt='EFFECTIVE',q='',sc='OBSERVED EVENT',attr='Tom Delaney Termination + vendor notice + complaint',act='ACT-0002',src='SRC-0105; SRC-0107',tag='TERM',hot='Y',corr='MULTIPLE INDEPENDENT',
   text="Tom Delaney's employment with Omni terminated, effective 2025-05-09. Confirmed by Veronica Leyva's 2025-05-16 vendor revocation notice (Tom 'No Longer with Omni (Effective May 9, 2025)') and pleaded in the complaint.",
   loc='Termination PDF (SRC-0105) is undated; effective date 2025-05-09 per Veronica vendor notice 2025-05-16 and complaint. Vendor notice not yet preserved as its own source.'),
 dict(id='EVT-0022',date='2025-04-03',dt='EVENT',q='',sc='SYSTEM RECORD',attr='Buildertrend Change Orders grid',act='ACT-0001',src='SRC-0109',tag='CHG',hot='',corr='SINGLE SOURCE',
   text='Change Order CO-0028 approved by David Wilson: deduct -$3,061.27.',loc='CO grid; TXN-0015. Approved by David Wilson 2025-04-03.'),
 dict(id='EVT-0023',date='2025-05-24',dt='EVENT',q='',sc='SYSTEM RECORD',attr='Buildertrend Change Orders grid',act='ACT-0001; ACT-0002',src='SRC-0109',tag='CHG',hot='Y',corr='SINGLE SOURCE',
   text="Change Order CO-0036 approved by David Wilson: -$63,917.18, 'Schneider to own the entire Ramada and finishes.' Created 2025-05-23, approved 2025-05-24, 15 days after Tom's 2025-05-09 termination. Ramada handed to Bradley Schneider (ORG-0008).",
   loc='CO grid; TXN-0013. Ramada handoff to Bradley Schneider (ORG-0008).'),
 dict(id='EVT-0024',date='2025-08-11',dt='EVENT',q='',sc='SYSTEM RECORD',attr='Buildertrend Change Orders grid',act='ACT-0001',src='SRC-0109',tag='CHG',hot='Y',corr='SINGLE SOURCE',
   text="Change Order CO-0048 approved by David Wilson: -$114,857.80, 'Remove Landscaping packages from Omni Scope.'",loc='CO grid; TXN-0014. Approved by David Wilson 2025-08-11.'),
 dict(id='EVT-0025',date='2025-08-29',dt='EVENT',q='',sc='SYSTEM RECORD',attr='Buildertrend Change Orders grid',act='ACT-0001',src='SRC-0109',tag='CHG',hot='',corr='SINGLE SOURCE',
   text='Change Order CO-0051 approved by David Wilson: deduct -$1,279.81.',loc='CO grid; TXN-0016. Approved by David Wilson 2025-08-29.'),
 dict(id='EVT-0026',date='2025-11-01',dt='EVENT',q='',sc='SYSTEM RECORD',attr='Buildertrend Invoice Payments grid',act='ACT-0001',src='SRC-0111',tag='PAY',hot='',corr='SINGLE SOURCE',
   text='Ninth and final recorded progress payment. 9 payments total $760,467.41 (first 2025-02-05, last 2025-11-01), all QuickBooks-synced. Invoice 0007 (5% punch-list retention, $51,752.62) remains unpaid.',
   loc='Payments grid; feeds R-0007. Full 9-payment breakdown in SRC-0111 notes and R-0007.'),
 dict(id='EVT-0027',date='2026-01-26',dt='FILED',q='on or about',sc='PARTY ASSERTION',attr='ROC complaint thread (Wilson / Heath Richards, AZ ROC)',act='ACT-0001',src='SRC-0117',tag='DEF',hot='Y',corr='SINGLE SOURCE',
   text="Wilson escalated the pool INTERIOR FINISH to the AZ Registrar of Contractors, complaint #2026-00144, pre-suit. By 2026-01-26 Wilson wrote 'We have reached a resolution regarding the interior finish of the pool'; complaint closed. Wilson later filed the civil suit 2026-06-12.",
   loc="martin@ ROC thread; complaint #2026-00144 re interior finish. Filing on or about 2026-01 (file ref '20260107'); resolution 2026-01-26."),
 dict(id='EVT-0028',date='2026-06-12',dt='FILED',q='',sc='DOCUMENT CONTENT',attr='Complaint (C20264565)',act='ACT-0001; ACT-0002; ACT-0003',src='SRC-0107',tag='DEM',hot='Y',corr='SINGLE SOURCE',
   text='David Wilson filed the civil complaint (Pima County Superior Court C20264565) against Omni, Tom and Bridget Delaney, and Sales One LLC. Counts: breach of contract, good faith and fair dealing, fraud/fraudulent inducement, negligent retention/supervision.',
   loc='Complaint C20264565. Filed 2026-06-12. Omni served 2026-07-02; answer due 2026-07-22.'),
 dict(id='EVT-0029',date='2026-07-16',dt='SENT',q='',sc='QUOTATION',attr='David Wilson (email)',act='ACT-0001; ACT-0009',src='SRC-0116',tag='WARR',hot='Y',corr='SINGLE SOURCE',
   text="Wilson replied to Allysa Redmon's compiled open-items list: 'I did spend time walking around and believe that is everything.' Confirmed the punch/warranty list complete (misters, grotto door, fiber-optic remote, steam sensor, travertine pavers sinking, dark grout).",
   loc="construction@ thread 'List of Open Items', Wilson reply 2026-07-16 16:54."),
 dict(id='EVT-0030',date='2026-07-19',dt='SENT',q='',sc='QUOTATION',attr='David Wilson (email)',act='ACT-0001; ACT-0005; ACT-0007; ACT-0006',src='SRC-0118',tag='TERM',hot='Y',corr='SINGLE SOURCE',
   text="Wilson emailed Michael Baker (forwarded to ownership): 'you, Scott and Martin came to my house and cornered me regarding Tom's termination and criminal behavior.' Wilson himself places the 2025-05-19 warning meeting and concedes Omni warned him about Tom; attributes defects to 'Tom's criminal and deviant behavior... corrected without Omni'; threatens another contractor over 'severe damage.'",
   loc='gmail thread 19f72526564ff643; forward msg 19f7b36f8054112b. Native .eml pending forensic re-export.'),
]
assert ws.cell(row=24,column=1).value=='EVT-0019'
def P(r,k,v): set_literal(ws.cell(row=r,column=c[k]),v)
for i,e in enumerate(EV):
    r=25+i
    assert not ws.cell(row=r,column=1).value, 'row %d not empty'%r
    P(r,'Event ID',e['id']); P(r,'Event Date',e['date']); P(r,'Date Type',e['dt'])
    if e['q']: P(r,'Qualifier',e['q'])
    P(r,'Event Text (released core)',e['text']); P(r,'Statement Class',e['sc']); P(r,'Attribution / Speaker',e['attr'])
    P(r,'Actors',e['act']); P(r,'Source ID(s)',e['src']); P(r,'Locator',e['loc']); P(r,'Issue Tags',e['tag'])
    if e['hot']: P(r,'Hot','Y')
    P(r,'Dispute Status','UNDISPUTED FOR NOW'); P(r,'Corroboration',e['corr'])
    for k,v in CONST.items(): P(r,k,v)
    P(r,'Batch / Ordinal','HAND-ENTRY-2026-07-20/%d'%(i+1))
    ws.cell(row=r,column=c['Check (auto)']).value='=IF($A%d="","",TRIM(IF($B%d="","NO DATE ","")&IF($J%d="","NO SOURCE ","")&IF($L%d="","NO TAGS ","")))'%(r,r,r,r)
    ws.cell(row=r,column=c['HotRank (auto)']).value='=IF($M%d="Y",COUNTIF($M$6:$M%d,"Y"),"")'%(r,r)
    print('added',e['id'],'row',r)
wb.save(WB); print('SAVED')
