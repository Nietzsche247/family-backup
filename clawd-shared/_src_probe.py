import sys
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map
from openpyxl import load_workbook
s=load_json_strict(r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1\Wilson.schema.json')
wb=load_workbook(r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1\Wilson.xlsx')
ws=wb['SOURCE INDEX']; cm=col_map(s,'SOURCE INDEX')
def g(r,k):
    v=ws.cell(row=r,column=cm[k]).value
    return str(v)[:26] if v is not None else '-'
nxt=None
for r in range(6, ws.max_row+2):
    sid=ws.cell(row=r,column=cm['Source ID']).value
    if not sid:
        nxt=r; break
    print('r%d %-9s|%-13s|doc=%-11s dt=%-9s basis=%-11s|sha=%-10s|wc=%-24s|disp=%s'%(
        r, str(sid), g(r,'Type'), g(r,'Doc Date'), g(r,'Date Type'), g(r,'Date Basis'),
        g(r,'SHA-256'), g(r,'Working Copy (filename)'), g(r,'Disposition')))
print('NEXT EMPTY ROW =', nxt)
for k in ['source_types','date_type','date_basis','source_disposition','acquisition_method']:
    if k in s.get('enums',{}): print('ENUM',k,'=',s['enums'][k])
