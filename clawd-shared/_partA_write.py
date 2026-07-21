import sys, os, shutil, hashlib
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map, set_literal
from openpyxl import load_workbook
M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'; WB=M+r'\Wilson.xlsx'; SC=M+r'\Wilson.schema.json'; KEY=M+r'\preserved\keydocs'
assert not os.path.exists(M+r'\~$Wilson.xlsx'), "OPEN in Excel"
def H(fn): return hashlib.sha256(open(os.path.join(KEY,fn),'rb').read()).hexdigest()
h118=H('SRC-0118_2026-07-19_Wilson-cornered.eml'); h121=H('SRC-0121_2026-07-17_Warranties-letter.eml')
h122=H('SRC-0122_2026-07-17_David-Wilson-warranty-position.pdf'); h123=H('SRC-0123_2026-07-20_payment-history-reply.eml')
bk=M+r'\Wilson_prewrite_backup_2026-07-21_PARTA.xlsx'; shutil.copy2(WB,bk); print('BACKUP',bk)
s=load_json_strict(SC); wb=load_workbook(WB); si=wb['SOURCE INDEX']; c=col_map(s,'SOURCE INDEX')
def P(row,k,v): set_literal(si.cell(row=row,column=c[k]),v)
EB='RFC 5322 Date header (sender-supplied; unverified)'
# finalize SRC-0118 (row 19)
assert si.cell(row=19,column=c['Source ID']).value=='SRC-0118'
P(19,'SHA-256',h118); P(19,'Date Basis',EB); P(19,'Disposition','PRESERVED')
P(19,'Native (relative URI)','gmail:thread/19f72526564ff643 msg/19f7b36f8054112b (native .eml)')
P(19,'Acquisition Method','Native RFC822 .eml downloaded from Gmail (aaron@ mailbox) 2026-07-21 via message "Download message"; Michael Baker forward of Wilson 2026-07-19 reply. Replaces the 2026-07-20 connector reconstruction.')
cur=si.cell(row=19,column=c['Notes']).value or ''
P(19,'Notes',(str(cur).rstrip()+' FINALIZED 2026-07-21: native .eml + canonical SHA-256 set (connector reconstruction replaced); Disposition RECEIVED -> PRESERVED.').strip())
print('SRC-0118 finalized, sha',h118[:12])
# next empty
r=None
for rr in range(6,si.max_row+2):
    if not si.cell(row=rr,column=c['Source ID']).value: r=rr; break
assert r==24,'expected row 24 got %s'%r
NEW={24:{'Source ID':'SRC-0121','SHA-256':h121,'Doc Date':'2026-07-17','Date Type':'SENT','Date Basis':EB,'Type':'Email',
 'Description':'Michael Baker 2026-07-17 letter to Wilson ("Please see attached document") transmitting Omni warranty/completion position (attachment David Wilson.pdf = SRC-0122)',
 'Custodian':'Aaron Baker (aaron mailbox)','From / Author':'Michael Baker <michael@omnipools.com>','To / Recipients':'David Wilson <davidwilson0@yahoo.com>',
 'Native (relative URI)':'gmail: "Warranties and job completion items" 2026-07-17','Working Copy (filename)':'SRC-0121_2026-07-17_Warranties-letter.eml',
 'Acquisition Method':'Native RFC822 .eml downloaded from Gmail (aaron@ mailbox) 2026-07-21 via "Download message".','Received Date':'2026-07-21','Disposition':'PRESERVED',
 'Notes':'Transmittal for Omni position statement (David Wilson.pdf = SRC-0122). Wilson replied 2026-07-19 (SRC-0118) that the letter "won\'t be completing or honoring the contractual items or warranty but then states you will be completing limited warranty items."'},
 25:{'Source ID':'SRC-0122','SHA-256':h122,'Doc Date':'2026-07-17','Date Type':'DOCUMENT','Date Basis':'Attached to and sent with the 2026-07-17 email (SRC-0121); PDF embedded date not separately verified.','Type':'Correspondence',
 'Description':'Omni warranty/completion position statement ("David Wilson.pdf") - declines full contractual/warranty completion, offers limited warranty; the payment-conditioned warranty position',
 'Custodian':'Omni Pool Builders / Michael Baker','From / Author':'Omni (Michael Baker)','To / Recipients':'David Wilson',
 'Native (relative URI)':'attachment of gmail "Warranties and job completion items" 2026-07-17 (SRC-0121)','Working Copy (filename)':'SRC-0122_2026-07-17_David-Wilson-warranty-position.pdf',
 'Acquisition Method':'PDF extracted from the native 2026-07-17 email (SRC-0121), downloaded from Gmail 2026-07-21.','Received Date':'2026-07-21','Disposition':'PRESERVED',
 'Notes':'Omni formal position. Per Wilson 2026-07-19 reply (SRC-0118): the letter states Omni "won\'t be completing or honoring the contractual items or warranty" but "will be completing limited warranty items." Ties to the payment-conditioned-warranty defense and STAGING item 4. Exact PDF text extractable if needed.'},
 26:{'Source ID':'SRC-0123','SHA-256':h123,'Doc Date':'2026-07-20','Date Type':'SENT','Date Basis':EB,'Type':'Email',
 'Description':'Michael Baker 2026-07-20 reply - final payment WAS released via BT 2026-03-16 (never opened) + requested by text 2026-04-27 + offer to re-release; 7 embedded screenshots',
 'Custodian':'Aaron Baker (aaron mailbox)','From / Author':'Michael Baker <michael@omnipools.com>','To / Recipients':'David Wilson <davidwilson0@yahoo.com>',
 'Native (relative URI)':'gmail: "Re: Warranties and job completion items" 2026-07-20','Working Copy (filename)':'SRC-0123_2026-07-20_payment-history-reply.eml',
 'Acquisition Method':'Native RFC822 .eml downloaded from Gmail (aaron@ mailbox) 2026-07-21 via "Download message". Includes 7 embedded image.png screenshots.','Received Date':'2026-07-21','Disposition':'PRESERVED',
 'Notes':'KEY REBUTTAL to Wilson 2026-07-19 "I have not received any remaining invoices." Michael: "final payment was requested a few times. 1. We released the final payment to you via BT March 16, 2026 - you never opened this invoice. 2. We also requested Final payment Via text on April 27th 2026. 3. I see a few other texts asking for final payment as well... I will release the final payment to you again." 7 embedded screenshots = the BT invoice + texts (proof). Sources STAGING items 1 (BT 2026-03-16), 2 (text 2026-04-27), 3 (other texts). Feeds R-0008.'}}
for rr,d in NEW.items():
    for k,v in d.items(): P(rr,k,v)
    print('added',d['Source ID'],'row',rr)
# STAGING updates
st=wb['STAGING']; sc=col_map(s,'STAGING')
def S(row,k,v): set_literal(st.cell(row=row,column=sc[k]),v)
S(6,'Rough Date','2026-03-16'); S(6,'What Document Proves It','SRC-0123 + embedded BT invoice screenshot: final payment invoice released via Buildertrend 2026-03-16; Wilson never opened it.'); S(6,'Status','FOUND')
S(7,'Rough Date','2026-04-27'); S(7,'What Document Proves It','SRC-0123 + embedded text screenshot: final payment requested via text 2026-04-27.'); S(7,'Status','FOUND')
S(8,'What Document Proves It','SRC-0123: "a few other texts asking for final payment"; specific dates TBD - pull the text thread with 714-269-6161.'); S(8,'Status','FOUND')
S(9,'Rough Date','2026-07-17'); S(9,'What Document Proves It','Warranty position = SRC-0122 (David Wilson.pdf): declines full warranty, offers limited warranty. Confirm the exact "RE-ESTABLISHED upon final payment" language (may be in SRC-0122 or an earlier email); Michael offers to re-release payment in SRC-0123.'); S(9,'Status','FOUND')
print('STAGING items 1-4 dated + FOUND')
# R-0008 append
rc=wb['RECONCILIATION']; rcm=col_map(s,'RECONCILIATION')
assert rc.cell(row=13,column=1).value=='R-0008'
cur8=rc.cell(row=13,column=rcm['Notes']).value or ''
set_literal(rc.cell(row=13,column=rcm['Notes']),(str(cur8).rstrip()+' | REBUTTAL (SRC-0123, Michael 2026-07-20): the final-payment invoice (0007, 5% retention) was RELEASED to Wilson via Buildertrend 2026-03-16 and Wilson never opened it; also requested by text 2026-04-27 and other texts; Michael offered to re-release. Directly rebuts Wilson 2026-07-19 "I have not received any remaining invoices."').strip())
print('R-0008 note updated')
wb.save(WB); print('SAVED')
