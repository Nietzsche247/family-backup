import sys, os, shutil
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map, set_literal
from openpyxl import load_workbook

M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'
WB=M+r'\Wilson.xlsx'; SC=M+r'\Wilson.schema.json'
assert not os.path.exists(M+r'\~$Wilson.xlsx'), "OPEN in Excel"
bk=M+r'\Wilson_prewrite_backup_2026-07-20_BATCH.xlsx'; shutil.copy2(WB,bk); print("BACKUP",bk)
s=load_json_strict(SC); wb=load_workbook(WB)
si=wb['SOURCE INDEX']; c=col_map(s,'SOURCE INDEX')
def put(ws,r,col,val): set_literal(ws.cell(row=r,column=c[col]) if ws is si else ws.cell(row=r,column=col), val)

EMAIL_BASIS='RFC 5322 Date header (sender-supplied; unverified)'
ACQ='A1 administrative export; Google Workspace Gmail API service-account impersonation under domain-wide delegation, read-only, 2026-07-17. Native RFC822 preserved from the corpus.'

# --- SRC-0115 (row16) fill hollow slot ---
assert si.cell(row=16,column=c['Source ID']).value=='SRC-0115'
put(si,16,'SHA-256','b742679a110490ea4c6f167a81c9d86a525b81052085cf4c9ac8fac4451533fd')
put(si,16,'Date Basis','NO DOCUMENT DATE - source bytes contain no accepted date metadata; Christine to confirm canonical/signed date')
put(si,16,'Native (relative URI)','dir:mail/christine.stewart/1980b0c01d3baa20.eml (attachment: Tom Cease and Desist.pdf)')
put(si,16,'Working Copy (filename)','SRC-0115_0000-00-00_Tom_Cease_and_Desist.pdf')
put(si,16,'Acquisition Method','A1 administrative export; Gmail API service-account impersonation, read-only, 2026-07-17. PDF extracted from native RFC822 message christine.stewart/1980b0c01d3baa20.eml ("Tom Delaney Documents", 2025-07-14).')
put(si,16,'Disposition','PRESERVED')
put(si,16,'Notes','PRESERVED 2026-07-20: working copy + SHA-256 set from the Tom Cease and Desist.pdf attachment to christine.stewart/1980b0c01d3baa20.eml. Christine to confirm this is the final/signed version or supply the signed original. Supports CD-0017 and CD-0019. SRC-0110..0114 reserved for pending invoice/payment exports.')
print("SRC-0115 filled")

# --- SRC-0116 (row17) Wilson 'that is everything' ---
assert not si.cell(row=17,column=c['Source ID']).value
rows={17:{'Source ID':'SRC-0116','SHA-256':'6088bce8331b056577ab03a9421ec62cecdc37ce545435222d3b4ffeaab86c26',
 'Doc Date':'2026-07-16','Date Type':'SENT','Date Basis':EMAIL_BASIS,'Type':'Email',
 'Description':'Re: List of Open Items - Wilson confirms open-items/warranty punch list is complete',
 'Custodian':'Allysa Redmon (construction@omnipoolbuilders.com)','From / Author':'David Wilson <davidwilson0@yahoo.com>',
 'To / Recipients':'Allysa Redmon <construction@omnipoolbuilders.com>','Native (relative URI)':'dir:mail/construction/19f6d5aa510117ba.eml',
 'Working Copy (filename)':'SRC-0116_2026-07-16_Re-List-of-Open-Items.eml','Acquisition Method':ACQ,'Received Date':'2026-07-17','Disposition':'PRESERVED',
 'Notes':'msgid 5B1F1A1F-7F20-49E3-8F88-C0125985EB9C@yahoo.com. Wilson: "I did spend time walking around and believe that is everything" - confirms the compiled open-items/warranty list is complete after Allysa asked him to review and add any missing items. Shield against post-site-visit escalation or later-added claims. Part of the 2026-07-16 "List of Open Items" thread (Allysa Redmon; Michael Baker looped).'},
 18:{'Source ID':'SRC-0117','SHA-256':'76d1f03e33878bceec5108828cb541ecbffe3815fddd9aeae501ee4788320cd2',
 'Doc Date':'2026-01-26','Date Type':'SENT','Date Basis':EMAIL_BASIS,'Type':'Email',
 'Description':'ROC complaint 2026-00144 thread - Wilson / Heath Richards (AZ Registrar of Contractors) re pool interior finish',
 'Custodian':'Martin Lopez (martin@omnipoolbuilders.com)','From / Author':'David Wilson <davidwilson0@yahoo.com>',
 'To / Recipients':'Heath Richards <heath.richards@roc.az.gov> (thread incl. Martin Lopez, Christine Stewart)','Native (relative URI)':'dir:mail/martin/19bfc505e9356124.eml',
 'Working Copy (filename)':'SRC-0117_2026-01-26_ROC-2026-00144.eml','Acquisition Method':ACQ,'Received Date':'2026-07-17','Disposition':'PRESERVED',
 'Notes':'msgid D48C4CC7-C5FB-4537-A045-0027E151D15C@yahoo.com. PRE-SUIT ESCALATION: Wilson filed AZ ROC complaint #2026-00144 re the pool INTERIOR FINISH (ref "20260107 Building Confidence Jobsite 2026-00144.pdf"). Wilson 2026-01-26: "We have reached a resolution regarding the interior finish of the pool"; complaint closed ~2026-01-26. Wilson nonetheless filed the civil suit 2026-06-12. Escalation channel separate from the warranty/open-items list.'}}
for r,d in rows.items():
    for k,v in d.items(): put(si,r,k,v)
    print("added",d['Source ID'])

# --- SRC-0102 note: flag Garrahan attachment ---
cur=si.cell(row=8,column=c['Notes']).value or ''
put(si,8,'Notes',(str(cur).rstrip()+' | OTHER-CLIENT DATA NOTE (added 2026-07-20): the native "Tom!!" message carried attachment "Garrahan _ REVISED DESIGN IMAGES - 10_23_23.pdf" (a different client, Garrahan) inadvertently included when Michael sent Wilson the Tom termination proof. Not broken out as a source. Flagged for counsel re inadvertent third-party disclosure.').strip())
print("SRC-0102 note flagged")

# --- CLAIMS notes ---
cl=wb['CLAIMS & DEFENSES']; cc=col_map(s,'CLAIMS & DEFENSES'); nc=cc['Notes']
assert cl.cell(row=23,column=cc['ID']).value=='CD-0018' and cl.cell(row=24,column=cc['ID']).value=='CD-0019'
add18=' MAY-2025 REMOVAL MECHANISM (per AB 2026-07-20): the ramada + landscaping were removed at the CLIENT direction, facilitated by Tom Delaney - Tom told Wilson he would manage those items and forwarded Wilson the DIRECT (pre-markup) sub bids Tom had obtained; Wilson, upset Omni had marked those items up, demanded they be removed so he could go direct. The removal COs (incl CO-0036 Ramada credit -$63,917.18, e-approved by David Wilson 2025-05-24, to Bradley Schneider) reflect Wilson choosing to self-source/go-direct, not an Omni failure to perform. Follow-up 2025-05-24 "Ramada Change" emails (corpus; Gemini source 139): Wilson asks Omni to help Bradley source ramada materials (blue tile, fan, lights) and handle logistics, confirms fireplace still Omni.'
add19=' WILSON RECOUNTS THE WARNING (weekend email ~2026-07-18/19, post-suit; NOT yet in preserved corpus, to be supplied): Wilson wrote that "Michael and Scott cornered him" - his own reference to the 2025-05-19 meeting, conceding Omni warned him about Tom. Meeting corroborated in "Christine & Ownership" internal chat (Gemini source 49): Michael Baker + Scott Culver, Wilson residence, 2025-05-19 9:30-10:00 AM. REMOVAL/GO-DIRECT: soon after the warning, Tom + Wilson worked together to strip the ramada + landscaping from Omni bid - Tom forwarded pre-markup direct bids and offered to manage the items; Wilson, angry about markup, demanded removal and went direct (see CD-0018). Wilson affirmatively engaging Tom/Schneider against Omni express warnings. TO PRESERVE: the weekend "cornered" email; Read AI/Gemini transcript of the 2025-05-19 meeting; "Christine & Ownership" + Managers chats (Gemini sources 49/37/73).'
set_literal(cl.cell(row=23,column=nc),(str(cl.cell(row=23,column=nc).value or '').rstrip()+add18).strip())
set_literal(cl.cell(row=24,column=nc),(str(cl.cell(row=24,column=nc).value or '').rstrip()+add19).strip())
print("CD-0018 + CD-0019 notes updated")
wb.save(WB); print("SAVED")
