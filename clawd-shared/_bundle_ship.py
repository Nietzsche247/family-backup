import os, sys, zipfile, hashlib
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map
from openpyxl import load_workbook
M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'; pres=M+r'\preserved'
STRAT='Wilson_Counsel_Strategy_FINAL_2026-07-21.xlsx'
def sha(p):
    h=hashlib.sha256()
    with open(p,'rb') as f:
        for c in iter(lambda:f.read(1024*1024),b''): h.update(c)
    return h.hexdigest()
s=load_json_strict(M+r'\Wilson.schema.json'); wb=load_workbook(M+r'\Wilson.xlsx'); si=wb['SOURCE INDEX']; c=col_map(s,'SOURCE INDEX')
L=['WILSON v OMNI POOL BUILDERS - EVIDENCE SOURCE MANIFEST','Case: Pima County Superior Court C20264565',
   'Generated: 2026-07-21  |  Working evidence set (attorney verification pending)','',
   'SOURCE ID | TYPE | SHA-256 (first 16) | WORKING COPY | DESCRIPTION','']
n=0
for r in range(6, si.max_row+1):
    sid=si.cell(row=r,column=c['Source ID']).value
    if not sid: continue
    n+=1
    L.append(' | '.join([str(sid), str(si.cell(row=r,column=c['Type']).value or '-'),
        (str(si.cell(row=r,column=c['SHA-256']).value or '')[:16] or '-'),
        str(si.cell(row=r,column=c['Working Copy (filename)']).value or '-'),
        str(si.cell(row=r,column=c['Description']).value or '')[:90]]))
L+=['','Total sources: %d'%n]
open(M+r'\SOURCE_MANIFEST.txt','w',encoding='utf-8').write('\n'.join(L))

HN="""WILSON v OMNI POOL BUILDERS - COUNSEL HANDOFF NOTE
Case: Pima County Superior Court C20264565
Prepared: 2026-07-21  |  Working evidence set - attorney verification required

DEADLINE (COUNSEL MUST INDEPENDENTLY CONFIRM IMMEDIATELY):
Complaint filed 2026-06-12; service reportedly 2026-07-02; answer reportedly due
2026-07-22. Counsel must independently confirm immediately. This package does not
calculate or rely on any deadline.

FULL VAULT CHAT CONTAINER:
Full SRC-0130 Google Vault container preserved separately and available through
secure transfer. (Approx. 1.39 GB; not embedded in this bundle due to size.
Google MD5 and internal SHA-256 are recorded in the workbook SOURCE INDEX;
SRC-0131 is the extracted warning-meeting subset, included here.)

WHAT THIS IS:
A clerical evidence-organization work product for counsel to verify. Every fact
links to a preserved, hash-verified source. Counsel Status is UNREVIEWED
throughout by design; this package makes no legal-sufficiency determinations.

CONTENTS:
- Wilson.xlsx                                    the evidence workbook
- Wilson_Counsel_Strategy_FINAL_2026-07-21.xlsx  strategy summary, generated
                                                 from the final workbook
- preserved/                                     all preserved source files
- SOURCE_MANIFEST.txt                            source list with SHA-256 hashes

TECHNICAL STATE OF THE WORKBOOK:
omni_audit 0 issues; omni_single_entry 0 issues; RECONCILIATION structure repaired
(header at row 8, R-0001 to R-0012 in rows 9-20, generated formulas restored);
native Microsoft Excel recalculation 0 formula errors. Superseded strategy files
(frozen 2026-07-17, v2, v3) are excluded from this bundle.

KNOWN OPEN ITEMS (flagged for counsel):
1. EVT-0039 (final-payment text on or about 2026-04-27) rests on Michael Baker's
   account in SRC-0123; the primary text screenshot is not yet preserved.
2. Two balance figures disagree: $51,752.62 (invoice grid) vs $36,377.04
   (job-remaining). The $15,375.58 delta is unreconciled; Christine Stewart to
   explain.
3. Workbook edits were direct-entry, not the hash-chained pipeline; change
   integrity rests on the external audits (omni_audit and omni_single_entry, both
   0 issues) plus dated backups. The sources themselves are all SHA-256 verified.
4. SRC-0112 (job-balance screenshot) not captured; the figure is documented in
   RECONCILIATION R-0008.
5. FACT REGISTRY is empty and DISPUTED FACTS remains a proposition list; DEADLINES
   is empty by design. Chronology records remain DRAFT and UNREVIEWED. Collection
   was targeted, not exhaustive.
"""
open(M+r'\HANDOFF_NOTE.txt','w',encoding='utf-8').write(HN)
BAD=['quarantine','backup','_prewrite','wilson_repaired',' - copy','strategy','.py','~$','vault_export','.tmp']
bundle=M+r'\WILSON_evidence_bundle_2026-07-21.zip'
if os.path.exists(bundle): os.remove(bundle)
skipped=[]
with zipfile.ZipFile(bundle,'w',zipfile.ZIP_DEFLATED) as z:
    for f in ['Wilson.xlsx','HANDOFF_NOTE.txt','SOURCE_MANIFEST.txt',STRAT]:
        z.write(os.path.join(M,f), f)
    for root,dirs,files in os.walk(pres):
        dirs[:]=[d for d in dirs if not any(b in d.lower() for b in BAD)]
        for f in files:
            fp=os.path.join(root,f); rel='preserved/'+os.path.relpath(fp,pres).replace('\\','/')
            if any(b in rel.lower() for b in BAD) or os.path.getsize(fp)>100*1024*1024:
                skipped.append(rel); continue
            z.write(fp,rel)
with zipfile.ZipFile(bundle) as z: names=z.namelist()
stale=[x for x in names if 'strategy' in x.lower() and x!=STRAT]
banned=[x for x in names if any(b in x.lower() for b in ['quarantine','backup','wilson_repaired',' - copy','.py'])]
print('=== FINAL COUNSEL BUNDLE ===')
print('filename        :', os.path.basename(bundle))
print('size            : %.2f MB'%(os.path.getsize(bundle)/1e6))
print('bundle SHA-256  :', sha(bundle))
print('exact file count:', len(names))
print('Wilson.xlsx SHA :', sha(M+r'\Wilson.xlsx'))
print('strategy SHA    :', sha(os.path.join(M,STRAT)))
print('sources         :', n)
print('stale strategy  :', stale if stale else 'NONE - confirmed removed')
print('banned artifacts:', banned if banned else 'NONE')
print('skipped by rule :', skipped if skipped else 'none')
print('\nroot-level entries:')
for x in names:
    if '/' not in x: print('   ',x)
