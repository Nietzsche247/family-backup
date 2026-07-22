import sys
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'
s=load_json_strict(M+r'\Wilson.schema.json'); wb=load_workbook(M+r'\Wilson.xlsx')
ws=wb['CLAIMS & DEFENSES']; c=col_map(s,'CLAIMS & DEFENSES')
print('CD columns:', {k:v for k,v in c.items()})
inv={v:k for k,v in c.items()}
# locate CD-0015/0018/0019
rows={}
for r in range(1,ws.max_row+1):
    a=ws.cell(row=r,column=1).value
    if isinstance(a,str) and a in ('CD-0015','CD-0018','CD-0019'): rows[a]=r
print('rows:',rows)
for cd in ['CD-0015','CD-0018','CD-0019']:
    r=rows[cd]; print('\n=== %s (row %d) ==='%(cd,r))
    for col in range(1,len(c)+2):
        v=ws.cell(row=r,column=col).value
        if v not in (None,''): print('  %-26s = %s'%(inv.get(col,'col%d'%col), str(v)[:200]))
