import sys, os, shutil
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map, set_literal
from openpyxl import load_workbook

M = r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'
WB = os.path.join(M, 'Wilson.xlsx'); SC = os.path.join(M, 'Wilson.schema.json')
assert not os.path.exists(os.path.join(M, '~$Wilson.xlsx')), "OPEN in Excel"
bk = os.path.join(M, 'Wilson_prewrite_backup_2026-07-20_CD19_SRC0115.xlsx')
shutil.copy2(WB, bk); print("BACKUP", bk)
schema = load_json_strict(SC); wb = load_workbook(WB)

# ---- CD-0019 defense (CLAIMS & DEFENSES row 24) ----
cl = wb['CLAIMS & DEFENSES']; cc = col_map(schema, 'CLAIMS & DEFENSES')
assert cl.cell(row=24, column=cc['ID']).value in (None,''), "CLAIMS row24 not empty"
cd19 = {
 'ID':'CD-0019','Side':'OURS',
 'Claim / Defense':"Plaintiff engaged Delaney independently against Omni's express warnings",
 'Count / Cite':"Defense to Counts One-Four (assumption of risk / superseding cause / failure to mitigate / comparative fault)",
 'Element or Predicate':("Omni met with Wilson and expressly warned him not to trust or deal with Delaney, banned Delaney from the job repeatedly, and issued a written cease-and-desist to Delaney. Despite this, Wilson - an out-of-state owner who was overseas during the key period - independently and informally engaged Delaney to manage projects across the property beyond Omni's scope (incl. a garage-floor overlay and bidding at property auctions on Wilson's behalf). Any damages flow from Wilson's own choice to engage Delaney against Omni's warnings, not from an Omni act or omission."),
 'Burden':'Omni','Status':'RESEARCH PENDING','Counsel Status':'UNREVIEWED',
 'Notes':("THEORY per AB account 2026-07-20; counsel to refine legal framing and elements. Evidence located/collected: "
   "(a) SRC-0103 Wilson 2025-09-11 'Tom stole over $95k from me... thank you for the help' - plaintiff blames Delaney and thanks Omni; "
   "(b) SRC-0109 CO grid incl CO-0036 Ramada handoff to Bradley Schneider; "
   "(c) Wilson 2025-05-24 'Re: Ramada Change' emails - Wilson coordinating Schneider on ramada logistics himself; "
   "(d) Wilson 2025-07-29 email to Wayne Parker / JE Pierre LLC re pool square footage - separate contractor; "
   "(e) Wilson 2026-01-26 emails to Heath Richards @roc.az.gov (ROC ref 2026-00144); "
   "(f) Wilson Las Vegas address + 2025 emails in +0900/+0800 (Asia) timezones = absentee/overseas owner; "
   "(g) SRC-0115 Tom Cease and Desist; (h) Veronica 2025-05-16 vendor revocation notice. "
   "TO COLLECT: in-person warning meeting record (daily logs/calendar); Wilson's 'why are you doing this to Tommy' message (comments/text); garage-overlay and auction-bidding specifics.")
}
for f, v in cd19.items():
    set_literal(cl.cell(row=24, column=cc[f]), v)
print("CD-0019 written (row 24)")

# ---- C&D slot SRC-0115 (SOURCE INDEX next empty row) ----
si = wb['SOURCE INDEX']; sc = col_map(schema, 'SOURCE INDEX'); idc = sc['Source ID']
r = 2
while si.cell(row=r, column=idc).value not in (None,''): r += 1
print("SOURCE INDEX next empty row:", r)
cd = {
 'Source ID':'SRC-0115','Type':'Correspondence',
 'Description':'Cease-and-desist letter issued by Omni to Tom Delaney (file: Tom Cease and Desist.pdf)',
 'Custodian':'Christine Stewart','From/Author':'Omni Pool Builders & Design LLC','To/Recipients':'Tom Delaney',
 'Disposition':'RECEIVED',
 'Notes':("SLOT for Christine Stewart to attach the canonical/signed Tom Cease and Desist.pdf and set SHA-256, Working Copy, "
   "Doc Date, and final Disposition. A copy is already located in the preserved email corpus as an attachment to "
   "Christine->Michael emails: 'Tom Delaney Documents' 2025-07-14 (christine.stewart/1980b0c01d3baa20.eml) and "
   "'Tom Delaney - Write Ups, Termination etc' 2026-07-13 (christine.stewart/19f5cbc7b9d1129e.eml). Christine to confirm "
   "this is the final version or supply the signed original. Supports CD-0017 and CD-0019. SRC-0110..0114 reserved for the pending invoice/payment exports.")
}
# use exact schema field names, tolerant of slash variants
def putcol(name_options, value):
    for nm in name_options:
        if nm in sc:
            set_literal(si.cell(row=r, column=sc[nm]), value); return
    raise KeyError(name_options)
putcol(['Source ID'],'SRC-0115')
putcol(['Type'],'Correspondence')
putcol(['Description'],cd['Description'])
putcol(['Custodian'],'Christine Stewart')
putcol(['From/Author','From / Author'],'Omni Pool Builders & Design LLC')
putcol(['To/Recipients','To / Recipients'],'Tom Delaney')
putcol(['Disposition'],'RECEIVED')
putcol(['Notes'],cd['Notes'])
print("SRC-0115 slot written (row %d)" % r)

wb.save(WB); print("SAVED")
