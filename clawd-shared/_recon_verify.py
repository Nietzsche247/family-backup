from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'
wb=load_workbook(M+r'\Wilson_repaired.xlsx')
ws=wb['RECONCILIATION']
print('tblRecon ref:', ws.tables['tblRecon'].ref)
print('table col names:', [tc.name for tc in ws.tables['tblRecon'].tableColumns])
print('\nrows 5-21 (F=formula):')
for r in range(5,22):
    cells=[]
    for col in range(1,16):
        v=ws.cell(row=r,column=col).value
        if v in (None,''): continue
        pref='=' if (isinstance(v,str) and v.startswith('=')) else ''
        cells.append('%s:%s%s'%(get_column_letter(col),pref,str(v)[:22]))
    print(' r%-3d %s'%(r, ' | '.join(cells) if cells else '(empty)'))
# confirm every data row 9-20 has all 6 formulas
print('\nformula presence check rows 9-20:')
bad=[]
for r in range(9,21):
    for col in (5,7,9,11,12,13):
        v=ws.cell(row=r,column=col).value
        if not (isinstance(v,str) and v.startswith('=')): bad.append((r,get_column_letter(col)))
print('  missing formulas:', bad if bad else 'NONE - all 12 rows have E/G/I/K/L/M')
