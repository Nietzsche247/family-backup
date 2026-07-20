import sys, os, shutil
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map, set_literal
from openpyxl import load_workbook

M = r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'
WB = os.path.join(M, 'Wilson.xlsx'); SC = os.path.join(M, 'Wilson.schema.json')
assert not os.path.exists(os.path.join(M, '~$Wilson.xlsx')), "OPEN in Excel"
bk = os.path.join(M, 'Wilson_prewrite_backup_2026-07-20_CLAIMS.xlsx')
shutil.copy2(WB, bk); print("BACKUP", bk)

schema = load_json_strict(SC); wb = load_workbook(WB)
ws = wb['CLAIMS & DEFENSES']; cm = col_map(schema, 'CLAIMS & DEFENSES')
Kstat = cm['Status']; Knote = cm['Notes']; Kid = cm['ID']

# 1) the three audit-flagged predicate_status cells -> COUNSEL REVIEW
for row, cid in [(7,'CD-0002'), (12,'CD-0007'), (15,'CD-0010')]:
    assert ws.cell(row=row, column=Kid).value == cid, f"{cid} not at row {row}"
    set_literal(ws.cell(row=row, column=Kstat), 'COUNSEL REVIEW')
print("CD-0002 / CD-0007 / CD-0010 Status -> COUNSEL REVIEW")

# 2) CD-0018 (row 23): evidence now collected this session
assert ws.cell(row=23, column=Kid).value == 'CD-0018', "CD-0018 not at row 23"
set_literal(ws.cell(row=23, column=Kstat), 'COUNSEL REVIEW')
cd18 = ('EVIDENCE NOW COLLECTED 2026-07-20: SRC-0109 (Buildertrend CO grid export) + transactions TXN-0013..0016. '
        '$183,116.06 of scope removed by client-approved change orders - CO-0036 -$63,917.18, CO-0048 -$114,857.80, '
        'CO-0028 -$3,061.27, CO-0051 -$1,279.81 - all e-approved by David Wilson (plaintiff) per the grid Client and '
        'Status columns. Predicate no longer missing; moved to COUNSEL REVIEW pending chronology-event linkage in '
        'the Supporting Event IDs column.')
set_literal(ws.cell(row=23, column=Knote), cd18)
print("CD-0018 updated: PREDICATE MISSING -> COUNSEL REVIEW, note reflects CO collection")

wb.save(WB); print("SAVED")
