import sys, os, email, shutil, hashlib
from email import policy
sys.path.insert(0, r'C:\Users\Aaron\.claude\skills\omni-litigation-engine\scripts')
from omni_core import load_json_strict, col_map
from openpyxl import load_workbook

M=r'C:\North_Star_Projects\WILSON_MATTER_WORKING_v1'
MAIL=r'C:\North_Star_Projects\Litigation-Wilson\preserved\mail'
KEY=os.path.join(M,'preserved','keydocs')
def sha(b): return hashlib.sha256(b).hexdigest()
def msgid(p):
    m=email.message_from_bytes(open(p,'rb').read(),policy=policy.default); return str(m.get('message-id',''))

# verify SRC-0101/0102/0103 mapping
s=load_json_strict(M+r'\Wilson.schema.json'); wb=load_workbook(M+r'\Wilson.xlsx')
ws=wb['SOURCE INDEX']; cm=col_map(s,'SOURCE INDEX')
for r in (7,8,9):
    print('SRC row',r,'desc=',ws.cell(row=r,column=cm['Description']).value,'| from=',ws.cell(row=r,column=cm['From / Author']).value,'| to=',ws.cell(row=r,column=cm['To / Recipients']).value)
print('---')

# 1) C&D PDF from christine.stewart/1980b0c01d3baa20.eml
src_cd=os.path.join(MAIL,'christine.stewart','1980b0c01d3baa20.eml')
m=email.message_from_bytes(open(src_cd,'rb').read(),policy=policy.default)
pdf=None
for part in m.walk():
    fn=(part.get_filename() or '')
    if 'cease' in fn.lower() and fn.lower().endswith('.pdf'):
        pdf=part.get_payload(decode=True); print('C&D attach found:',fn,len(pdf),'bytes'); break
if pdf:
    dst=os.path.join(KEY,'SRC-0115_0000-00-00_Tom_Cease_and_Desist.pdf')
    open(dst,'wb').write(pdf); print('SRC-0115 pdf SHA=',sha(pdf))
else: print('!! C&D PDF NOT FOUND')

# 2) Wilson 'that's everything' email
src16=os.path.join(MAIL,'construction','19f6d5aa510117ba.eml')
dst16=os.path.join(KEY,'SRC-0116_2026-07-16_Re-List-of-Open-Items.eml')
shutil.copy2(src16,dst16); print('SRC-0116 eml SHA=',sha(open(src16,"rb").read()),'| msgid=',msgid(src16))

# 3) ROC email
src17=os.path.join(MAIL,'martin','19bfc505e9356124.eml')
dst17=os.path.join(KEY,'SRC-0117_2026-01-26_ROC-2026-00144.eml')
shutil.copy2(src17,dst17); print('SRC-0117 eml SHA=',sha(open(src17,"rb").read()),'| msgid=',msgid(src17))
print('DONE extract')
